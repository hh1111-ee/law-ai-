import atexit
import json
import logging
import os
import sys
from logging.handlers import RotatingFileHandler
from typing import Dict, List, Optional, Tuple

import requests
from fastapi import FastAPI, Request, WebSocket, WebSocketDisconnect
from fastapi.middleware.cors import CORSMiddleware
from fastapi.responses import JSONResponse, PlainTextResponse
from openpyxl import load_workbook

BASE_DIR = os.path.abspath(os.path.dirname(__file__))
CHAT_BACKEND_DIR = os.path.join(BASE_DIR, "聊天和用户后端")

if CHAT_BACKEND_DIR not in sys.path:
    sys.path.append(CHAT_BACKEND_DIR)

from ChatMessage import MessageManage as MsgMgr  # noqa: E402
from ChatMessage import personalChatMessage  # noqa: E402
from group import groupManage as GroupMgr  # noqa: E402
from post import PostManage  # noqa: E402
from user import user as UserClass  # noqa: E402
from user import userManage as UserMgr  # noqa: E402

# ===================== 日志配置 =====================
LOG_DIR = os.path.join(BASE_DIR, "logs")
os.makedirs(LOG_DIR, exist_ok=True)
LOG_FORMAT = "%(asctime)s - %(levelname)s - %(module)s:%(funcName)s:%(lineno)d - %(message)s"
DATE_FORMAT = "%Y-%m-%d %H:%M:%S"

file_handler = RotatingFileHandler(
    os.path.join(LOG_DIR, "combined_server.log"),
    maxBytes=5 * 1024 * 1024,
    backupCount=5,
    encoding="utf-8",
)
file_handler.setFormatter(logging.Formatter(LOG_FORMAT, DATE_FORMAT))

console_handler = logging.StreamHandler()
console_handler.setFormatter(logging.Formatter(LOG_FORMAT, DATE_FORMAT))

logging.basicConfig(level=logging.INFO, handlers=[file_handler, console_handler])
logger = logging.getLogger(__name__)

# ===================== 数据路径 =====================
USER_FILE = os.path.join(BASE_DIR, "users.pkl")
GROUP_FILE = os.path.join(BASE_DIR, "groups.pkl")
POST_FILE = os.path.join(BASE_DIR, "posts.pkl")
PERSONAL_MSG_FILE = os.path.join(BASE_DIR, "personal_messages.pkl")
GROUP_MSG_FILE = os.path.join(BASE_DIR, "group_messages.pkl")

DATA_PATH = os.path.join(BASE_DIR, "数据", "地区.xlsx")

# ===================== 业务对象 =====================
app = FastAPI()
msg_manager = MsgMgr()
user_manager = UserMgr()
group_manager = GroupMgr()
post_manager = PostManage()

app.add_middleware(
    CORSMiddleware,
    allow_origins=["*"],
    allow_credentials=True,
    allow_methods=["*"],
    allow_headers=["*"],
)

# ===================== 工具函数 =====================
_DATA_CACHE: List[Dict[str, str]] = []
_DATA_MTIME: float = -1.0
_save_exit_called = False


def ensure_file_exists(filename: str) -> None:
    if not os.path.exists(filename):
        with open(filename, "wb"):
            pass
        logger.info("创建空数据文件：%s", filename)


def return_error(message: str, code: int = 400, request_data: Optional[Dict] = None) -> JSONResponse:
    error_info = {
        "code": code,
        "error": message,
        "request_data": request_data or {},
    }
    logger.error("接口错误：%s，请求参数：%s", message, error_info["request_data"])
    return JSONResponse(status_code=code, content=error_info)


def return_success(data: Optional[Dict] = None, message: str = "操作成功") -> JSONResponse:
    success_info = {
        "code": 200,
        "message": message,
        "data": data or {},
    }
    logger.info("接口成功：%s，返回数据：%s", message, data)
    return JSONResponse(content=success_info)


def load_all_data_on_start() -> None:
    ensure_file_exists(PERSONAL_MSG_FILE)
    ensure_file_exists(GROUP_MSG_FILE)
    ensure_file_exists(USER_FILE)
    ensure_file_exists(GROUP_FILE)
    ensure_file_exists(POST_FILE)

    msg_manager.load_personal_messages(PERSONAL_MSG_FILE)
    msg_manager.load_group_messages(GROUP_MSG_FILE)
    user_manager.load_users(USER_FILE)
    group_manager.load_groups(GROUP_FILE)
    post_manager.load_posts(POST_FILE)
    for user in user_manager.user_list:
        logger.info(
            "加载用户：%s，状态：%s，好友数：%s",
            user.username,
            user.state,
            len(user.get_friends()),
        )
    logger.info("所有数据加载成功")


def save_all_data_on_exit() -> None:
    global _save_exit_called
    if _save_exit_called:
        logger.warning("save_all_data_on_exit 已被调用，跳过重复保存")
        return
    _save_exit_called = True
    logger.info("开始保存所有数据（退出钩子）")
    msg_manager.save_personal_messages(PERSONAL_MSG_FILE)
    logger.info("个人消息保存完成")
    msg_manager.save_group_messages(GROUP_MSG_FILE)
    logger.info("群消息保存完成")
    user_manager.save_users(USER_FILE)
    logger.info("用户数据保存完成")
    group_manager.save_groups(GROUP_FILE)
    logger.info("群组数据保存完成")
    post_manager.save_posts(POST_FILE)
    logger.info("帖子数据保存完成")
    logger.info("所有数据保存成功")


def _safe_text(value) -> str:
    if value is None:
        return ""
    return str(value).strip()


def _load_excel_rows() -> List[Dict[str, str]]:
    if not os.path.exists(DATA_PATH):
        logger.warning("数据文件不存在: %s", DATA_PATH)
        return []

    workbook = load_workbook(DATA_PATH, read_only=True, data_only=True)
    sheet = workbook.active
    rows: List[Dict[str, str]] = []

    for row in sheet.iter_rows(min_row=2, max_col=3, values_only=True):
        region = _safe_text(row[0])
        url = _safe_text(row[1])
        law_name = _safe_text(row[2])
        if not (region or url or law_name):
            continue
        rows.append({"region": region, "url": url, "name": law_name})

    return rows


def get_data_rows() -> List[Dict[str, str]]:
    global _DATA_CACHE, _DATA_MTIME
    try:
        mtime = os.path.getmtime(DATA_PATH)
    except OSError:
        mtime = -1.0

    if mtime != _DATA_MTIME:
        _DATA_CACHE = _load_excel_rows()
        _DATA_MTIME = mtime
        logger.info("已加载数据: %s 条", len(_DATA_CACHE))

    return _DATA_CACHE


def resolve_location(payload: Dict) -> Dict[str, str]:
    location = payload.get("location") or {}
    province = _safe_text(payload.get("province") or location.get("province"))
    city = _safe_text(payload.get("city") or location.get("city"))
    region = _safe_text(payload.get("region"))

    if region and not province:
        province = region

    if not province:
        province = "全国"

    return {"province": province, "city": city or province}


def build_results(rows: List[Dict[str, str]], keyword: str) -> List[Dict[str, str]]:
    keyword_lower = keyword.lower()
    results: List[Dict[str, str]] = []

    for row in rows:
        name = row.get("name", "")
        region = row.get("region", "")
        if keyword_lower not in name.lower() and keyword_lower not in region.lower():
            continue
        results.append({
            "name": name or "未命名法律",
            "desc": f"地区: {region or '未知'}",
            "url": row.get("url", ""),
        })

    return results


def paginate_results(results: List[Dict[str, str]], page: int, page_size: int) -> Dict[str, object]:
    page = max(page, 1)
    page_size = max(min(page_size, 50), 1)
    total = len(results)
    start = (page - 1) * page_size
    end = start + page_size
    return {
        "page": page,
        "page_size": page_size,
        "total": total,
        "results": results[start:end],
    }


def build_recommended(rows: List[Dict[str, str]], location: Dict[str, str]) -> List[Dict[str, str]]:
    province = location.get("province", "")
    city = location.get("city", "")

    if province and province != "全国":
        matched = [
            row
            for row in rows
            if province in row.get("region", "") or city in row.get("region", "")
        ]
    else:
        matched = rows

    recommended: List[Dict[str, str]] = []
    seen = set()
    for row in matched:
        key = (row.get("name", ""), row.get("url", ""))
        if key in seen:
            continue
        seen.add(key)
        recommended.append({
            "title": row.get("name", "未命名法律"),
            "desc": f"来自: {row.get('region', '') or '未知'}",
        })
        if len(recommended) >= 6:
            break

    return recommended


def send_personal_message_logic(sender_name: str, receiver_name: str, content: str, timestamp: str) -> Tuple[bool, str]:
    sender = user_manager.find_user(sender_name)
    receiver = user_manager.find_user(receiver_name)

    if not sender:
        return False, f"发送方「{sender_name}」不存在"
    if not receiver:
        return False, f"接收方「{receiver_name}」不存在"

    message = personalChatMessage(sender, receiver, content, timestamp)
    msg_manager.add_personal_message(message)
    try:
        msg_manager.save_personal_messages(PERSONAL_MSG_FILE)
        logger.info("发送私信后个人消息已保存：%s", PERSONAL_MSG_FILE)
    except Exception as exc:
        logger.error("发送私信后保存个人消息失败：%s", exc, exc_info=True)

    return True, "私信发送成功"


class ConnectionManager:
    def __init__(self) -> None:
        self.active_connections: Dict[str, WebSocket] = {}

    async def connect(self, user_id: str, websocket: WebSocket) -> None:
        await websocket.accept()
        self.active_connections[user_id] = websocket
        logger.info("用户 %s 建立WebSocket连接，当前活跃连接数：%s", user_id, len(self.active_connections))

    def disconnect(self, user_id: str) -> None:
        if user_id in self.active_connections:
            self.active_connections.pop(user_id, None)
            logger.info("用户 %s 断开WebSocket连接，当前活跃连接数：%s", user_id, len(self.active_connections))

    async def send_to_user(self, user_id: str, message: Dict) -> None:
        websocket = self.active_connections.get(user_id)
        if not websocket:
            logger.warning("用户 %s 无活跃WebSocket连接，消息发送失败：%s", user_id, message)
            return
        await websocket.send_text(json.dumps(message, ensure_ascii=False))
        logger.info("向用户 %s 发送消息：%s", user_id, message)


manager = ConnectionManager()


async def _get_payload(request: Request) -> Optional[Dict]:
    try:
        return await request.json()
    except Exception:
        return None


@app.on_event("startup")
async def _startup() -> None:
    load_all_data_on_start()


_atexit_registered = False
if not _atexit_registered:
    atexit.register(save_all_data_on_exit)
    _atexit_registered = True


@app.exception_handler(Exception)
async def handle_global_exception(request: Request, exc: Exception):
    error_msg = f"服务器内部错误：{exc}"
    logger.error(error_msg, exc_info=True)
    payload = await _get_payload(request) or dict(request.query_params)
    return JSONResponse(
        status_code=500,
        content={"code": 500, "error": error_msg, "request_data": payload},
    )


@app.get("/")
async def root():
    return PlainTextResponse("Combined server is running.")


@app.post("/load_all_data")
async def load_all_data():
    try:
        load_all_data_on_start()
        return return_success(message="所有数据加载成功")
    except Exception as exc:
        return return_error(f"数据加载失败：{exc}", 500)


@app.post("/save_all_data")
async def save_all_data():
    try:
        save_all_data_on_exit()
        return return_success(message="所有数据保存成功")
    except Exception as exc:
        return return_error(f"数据保存失败：{exc}", 500)


@app.post("/user_register")
async def user_register(request: Request):
    data = await _get_payload(request)
    if not data:
        return return_error("请求参数不能为空，请传入JSON格式数据")

    username = data.get("username")
    identity = data.get("identity")
    password = data.get("password")
    location = data.get("location")
    role = data.get("role") or identity

    missing_fields = []
    if not username:
        missing_fields.append("用户名(username)")
    if not identity:
        missing_fields.append("身份标识(identity)")
    if not password:
        missing_fields.append("密码(password)")
    if not location:
        missing_fields.append("位置(location)")

    if missing_fields:
        return return_error(f"注册失败：缺少必填参数 - {', '.join(missing_fields)}")

    if user_manager.find_user(username):
        return return_error(f"注册失败：用户名「{username}」已存在", 400)

    try:
        new_user = UserClass(username, identity, password, location, role)
        user_manager.add_user(new_user)
        logger.info("用户注册成功：%s，身份：%s，位置：%s", username, identity, location)
        try:
            user_manager.save_users(USER_FILE)
            logger.info("注册后用户数据已保存：%s", USER_FILE)
        except Exception as exc:
            logger.error("注册后保存用户数据失败：%s", exc, exc_info=True)
        return return_success(data={"user": new_user.get_profile()}, message=f"用户「{username}」注册成功")
    except Exception as exc:
        return return_error(f"用户「{username}」注册失败：{exc}", 500)


@app.post("/user_login")
async def user_login(request: Request):
    data = await _get_payload(request)
    if not data:
        return return_error("请求参数不能为空，请传入JSON格式数据")

    username = data.get("username")
    password = data.get("password")

    if not username:
        return return_error("登录失败：缺少用户名(username)参数")
    if not password:
        return return_error("登录失败：缺少密码(password)参数")

    user_obj = user_manager.find_user(username)
    if not user_obj:
        return return_error(f"登录失败：用户名「{username}」不存在", 401)

    if user_obj.password != password:
        return return_error(f"登录失败：用户名「{username}」的密码错误", 401)

    user_obj.set_state("online")
    profile = user_obj.get_profile()
    profile.pop("password", None)
    return return_success(data={"user": profile}, message=f"用户「{username}」登录成功")


@app.post("/user_friends")
async def user_friends(request: Request):
    data = await _get_payload(request)
    if not data:
        return return_error("请求参数不能为空，请传入JSON格式数据")

    username = data.get("username")
    if not username:
        return return_error("查询失败：缺少用户名(username)参数")

    user_obj = user_manager.find_user(username)
    if not user_obj:
        return return_error(f"查询失败：用户名「{username}」不存在", 404)

    friends = [friend.get_profile() for friend in user_obj.get_friends()]
    return return_success(data={"friends": friends}, message=f"查询到用户「{username}」的好友列表")


@app.post("/user_state_search")
async def user_state_search(request: Request):
    data = await _get_payload(request)
    if not data:
        return return_error("请求参数不能为空，请传入JSON格式数据")

    username = data.get("username")
    if not username:
        return return_error("查询失败：缺少用户名(username)参数")

    user_in_state = [u.get_profile() for u in user_manager.user_list if u.username == username]
    if not user_in_state:
        return return_error(f"查询失败：未找到用户名「{username}」的用户信息", 404)

    return return_success(data={"users": user_in_state}, message=f"查询到用户「{username}」的状态信息")


@app.post("/send_personal_message")
async def send_personal_message(request: Request):
    data = await _get_payload(request)
    if not data:
        return return_error("请求参数不能为空，请传入JSON格式数据")

    sender_name = data.get("sender")
    receiver_name = data.get("receiver")
    content = data.get("content")
    timestamp = data.get("timestamp")

    missing_fields = []
    if not sender_name:
        missing_fields.append("发送方(sender)")
    if not receiver_name:
        missing_fields.append("接收方(receiver)")
    if not content:
        missing_fields.append("消息内容(content)")
    if not timestamp:
        missing_fields.append("时间戳(timestamp)")

    if missing_fields:
        return return_error(f"发送私信失败：缺少必填参数 - {', '.join(missing_fields)}")

    ok, message = send_personal_message_logic(sender_name, receiver_name, content, timestamp)
    if not ok:
        return return_error(f"发送私信失败：{message}", 404)

    return return_success(message=f"私信发送成功：从「{sender_name}」到「{receiver_name}」，内容：{content}")


@app.post("/send_group_message")
async def send_group_message(request: Request):
    data = await _get_payload(request)
    if not data:
        return return_error("请求参数不能为空，请传入JSON格式数据")

    sender_name = data.get("sender")
    group_name = data.get("group")
    content = data.get("content")
    timestamp = data.get("timestamp")

    missing_fields = []
    if not sender_name:
        missing_fields.append("发送方(sender)")
    if not group_name:
        missing_fields.append("群组名(group)")
    if not content:
        missing_fields.append("消息内容(content)")
    if not timestamp:
        missing_fields.append("时间戳(timestamp)")

    if missing_fields:
        return return_error(f"发送群消息失败：缺少必填参数 - {', '.join(missing_fields)}")

    sender = user_manager.find_user(sender_name)
    group = group_manager.find_group(group_name)

    error_msg = []
    if not sender:
        error_msg.append(f"发送方「{sender_name}」不存在")
    if not group:
        error_msg.append(f"群组「{group_name}」不存在")

    if error_msg:
        return return_error(f"发送群消息失败：{'; '.join(error_msg)}", 404)

    try:
        from ChatMessage import groupChatMessage

        message = groupChatMessage(sender, group, content, timestamp)
        msg_manager.add_group_message(message)
        return return_success(message=f"群消息发送成功：「{sender_name}」发送到群组「{group_name}」")
    except Exception as exc:
        return return_error(f"发送群消息失败：{exc}", 500)


@app.post("/get_personal_messages")
async def get_personal_messages(request: Request):
    data = await _get_payload(request)
    if not data:
        return return_error("请求参数不能为空，请传入JSON格式数据")
    user1 = data.get("user1")
    user2 = data.get("user2")
    if not user1 or not user2:
        return return_error("查询失败：缺少 user1 或 user2 参数")

    u1 = user_manager.find_user(user1)
    u2 = user_manager.find_user(user2)
    if not u1 or not u2:
        return return_error(f"查询失败：用户不存在（{user1 if not u1 else ''} {user2 if not u2 else ''})", 404)

    try:
        messages = msg_manager.get_personal_messages(user1, user2)

        def to_str(val):
            return getattr(val, "username", str(val))

        msg_list = []
        for msg in messages:
            msg_list.append(
                {
                    "sender": to_str(getattr(msg, "sender", getattr(msg, "from_user", ""))),
                    "receiver": to_str(getattr(msg, "receiver", getattr(msg, "to_user", ""))),
                    "content": getattr(msg, "content", getattr(msg, "text", "")),
                    "timestamp": getattr(msg, "timestamp", getattr(msg, "time", "")),
                }
            )
        return return_success(data={"messages": msg_list}, message=f"查询到{user1}与{user2}的历史私聊消息，共{len(msg_list)}条")
    except Exception as exc:
        logger.error("查询历史私聊消息失败：%s", exc, exc_info=True)
        return return_error(f"查询历史私聊消息失败：{exc}", 500)


@app.get("/get_posts")
async def get_posts(request: Request):
    section = request.query_params.get("section")
    if not section:
        return return_error("查询帖子失败：缺少板块名称(section)参数", request_data=dict(request.query_params))

    try:
        posts = post_manager.get_posts(section)
        posts_dict = [p.to_dict() for p in posts]
        return return_success(
            data={"posts": posts_dict},
            message=f"查询到「{section}」板块下共{len(posts_dict)}条帖子",
        )
    except Exception as exc:
        return return_error(f"查询「{section}」板块帖子失败：{exc}", 500, request_data=dict(request.query_params))


@app.post("/create_post")
async def create_post(request: Request):
    data = await _get_payload(request)
    if not data:
        return return_error("请求参数不能为空，请传入JSON格式数据")

    author = data.get("author")
    title = data.get("title")
    content = data.get("content")
    section = data.get("section")

    missing_fields = []
    if not author:
        missing_fields.append("作者(author)")
    if not title:
        missing_fields.append("标题(title)")
    if not content:
        missing_fields.append("内容(content)")
    if not section:
        missing_fields.append("板块(section)")

    if missing_fields:
        return return_error(f"创建帖子失败：缺少必填参数 - {', '.join(missing_fields)}")

    if not user_manager.find_user(author):
        return return_error(f"创建帖子失败：作者「{author}」不存在", 404)

    try:
        post = post_manager.add_post(author, title, content, section)
        try:
            post_manager.save_posts(POST_FILE)
            logger.info("创建帖子后帖子数据已保存：%s", POST_FILE)
        except Exception as exc:
            logger.error("创建帖子后保存帖子数据失败：%s", exc, exc_info=True)
        return return_success(
            data={"post": post.to_dict()},
            message=f"帖子「{title}」创建成功（ID：{post.id}）",
        )
    except Exception as exc:
        return return_error(f"创建帖子「{title}」失败：{exc}", 500)


@app.get("/get_post_detail")
async def get_post_detail(request: Request):
    post_id_str = request.query_params.get("post_id")
    if not post_id_str:
        return return_error("查询帖子详情失败：缺少帖子ID(post_id)参数", request_data=dict(request.query_params))

    try:
        post_id = int(post_id_str)
    except ValueError:
        return return_error(f"查询帖子详情失败：帖子ID「{post_id_str}」不是有效的数字")

    post = post_manager.get_post(post_id)
    if not post:
        return return_error(f"查询帖子详情失败：未找到ID为{post_id}的帖子", 404)

    return return_success(data={"post": post.to_dict()}, message=f"查询到ID为{post_id}的帖子详情")


@app.post("/add_comment")
async def add_comment(request: Request):
    data = await _get_payload(request)
    if not data:
        return return_error("请求参数不能为空，请传入JSON格式数据")

    post_id_str = data.get("post_id")
    author = data.get("author")
    content = data.get("content")

    missing_fields = []
    if not post_id_str:
        missing_fields.append("帖子ID(post_id)")
    if not author:
        missing_fields.append("评论作者(author)")
    if not content:
        missing_fields.append("评论内容(content)")

    if missing_fields:
        return return_error(f"添加评论失败：缺少必填参数 - {', '.join(missing_fields)}")

    try:
        post_id = int(post_id_str)
    except ValueError:
        return return_error(f"添加评论失败：帖子ID「{post_id_str}」不是有效的数字")

    if not user_manager.find_user(author):
        return return_error(f"添加评论失败：作者「{author}」不存在", 404)

    comment = post_manager.add_comment(post_id, author, content)
    if not comment:
        return return_error(f"添加评论失败：未找到ID为{post_id}的帖子", 404)

    return return_success(
        data={"comment": comment.to_dict()},
        message=f"成功为ID{post_id}的帖子添加评论（评论作者：{author}）",
    )


@app.post("/user_logout")
async def user_logout(request: Request):
    data = await _get_payload(request)
    if not data:
        return return_error("请求参数不能为空，请传入JSON格式数据")

    username = data.get("username")
    if not username:
        return return_error("登出失败：缺少用户名(username)参数")

    user_obj = user_manager.find_user(username)
    if not user_obj:
        return return_error(f"登出失败：用户名「{username}」不存在", 404)

    user_obj.set_state("offline")
    return return_success(message=f"用户「{username}」已成功登出，状态更新为离线")


@app.post("/add_friend")
async def add_friend(request: Request):
    data = await _get_payload(request)
    if not data:
        return return_error("请求参数不能为空，请传入JSON格式数据")

    username = data.get("username")
    friend_name = data.get("friend")

    if not username:
        return return_error("添加好友失败：缺少用户名(username)参数")
    if not friend_name:
        return return_error("添加好友失败：缺少好友用户名(friend)参数")

    if username == friend_name:
        return return_error(f"添加好友失败：不能添加自己（用户名：{username}）", 400)

    user_obj = user_manager.find_user(username)
    friend_obj = user_manager.find_user(friend_name)

    error_msg = []
    if not user_obj:
        error_msg.append(f"当前用户「{username}」不存在")
    if not friend_obj:
        error_msg.append(f"好友「{friend_name}」不存在")

    if error_msg:
        return return_error(f"添加好友失败：{'; '.join(error_msg)}", 404)

    if friend_obj in user_obj.get_friends():
        return return_error(f"添加好友失败：「{username}」已添加「{friend_name}」为好友", 400)

    try:
        user_obj.add_friend(friend_obj)
        friend_obj.add_friend(user_obj)
        return return_success(message=f"添加好友成功：「{username}」和「{friend_name}」已互加为好友")
    except Exception as exc:
        return return_error(f"添加好友失败：{exc}", 500)


@app.post("/search")
async def api_search(request: Request):
    payload = await _get_payload(request) or {}
    keyword = _safe_text(payload.get("keyword"))
    if not keyword:
        return return_error("请提供搜索关键字", 400)

    rows = get_data_rows()
    location = resolve_location(payload)
    all_results = build_results(rows, keyword)
    recommended = build_recommended(rows, location)

    try:
        page = int(payload.get("page", 1))
    except (TypeError, ValueError):
        page = 1
    try:
        page_size = int(payload.get("page_size", 5))
    except (TypeError, ValueError):
        page_size = 5

    page_data = paginate_results(all_results, page, page_size)

    return JSONResponse(
        content={
            "location": location,
            "recommended": recommended,
            "results": page_data["results"],
            "page": page_data["page"],
            "page_size": page_data["page_size"],
            "total": page_data["total"],
        }
    )


@app.post("/location")
@app.get("/location")
async def api_location(request: Request):
    if request.method == "GET":
        payload = {}
    else:
        payload = await _get_payload(request) or {}

    location = resolve_location(payload)
    return JSONResponse(content={"location": location})


@app.post("/api/legal")
async def legal_chat(request: Request):
    data = await _get_payload(request)
    if not data or "question" not in data:
        return return_error("请输入要咨询的法律问题", 400)

    question = str(data.get("question", "")).strip()
    if not question:
        return return_error("问题不能为空", 400)

    payload = {
        "model": "gpt-oss:120b-cloud",
        "stream": False,
        "messages": [
            {
                "role": "system",
                "content": "你是一名中华人民共和国执业律师，请在回答中**必须**引用具体的法条编号（如《民法典》第23条）。\n如果法律里没有对应条文，请直接回复“暂无相关法条”。\n请使用简洁、正式的法律语言，切勿捏造法条内容。",
            },
            {"role": "user", "content": question},
        ],
    }

    try:
        logger.info("→ 向 Ollama 发送请求：%s...", question[:30])
        resp = requests.post("http://127.0.0.1:11434/api/chat", json=payload, timeout=30)
        resp.raise_for_status()
        answer = resp.json().get("message", {}).get("content", "暂无相关法条")
        logger.info("✅ 得到答案（前50字符）: %s", answer[:50])
        return JSONResponse(content={"answer": answer})
    except requests.exceptions.ConnectionError:
        logger.exception("❌ 无法连接到 Ollama")
        return return_error("无法连接到 Ollama，请先启动 Ollama", 500)
    except Exception as exc:
        logger.exception("❌ 未知错误")
        return return_error(f"服务器内部错误：{exc}", 500)


@app.websocket("/ws/private")
async def websocket_private_chat(websocket: WebSocket):
    user_id = websocket.query_params.get("user_id")
    if not user_id:
        logger.error("WebSocket连接失败：未提供user_id参数，关闭连接")
        await websocket.close(code=1008)
        return

    await manager.connect(user_id, websocket)

    try:
        while True:
            data = await websocket.receive_text()
            logger.info("收到用户 %s 发送的原始数据：%s", user_id, data)

            try:
                payload = json.loads(data)
            except json.JSONDecodeError as exc:
                logger.error("用户 %s JSON解析失败：%s", user_id, exc)
                await manager.send_to_user(user_id, {"type": "error", "message": "Invalid JSON payload."})
                continue

            sender = payload.get("from")
            receiver = payload.get("to")
            content = payload.get("content")
            timestamp = payload.get("ts")

            missing_fields = []
            if not sender:
                missing_fields.append("from")
            if not receiver:
                missing_fields.append("to")
            if not content:
                missing_fields.append("content")

            if missing_fields:
                error_msg = f"缺失必要字段：{', '.join(missing_fields)}"
                logger.error("用户 %s %s，payload：%s", user_id, error_msg, payload)
                await manager.send_to_user(user_id, {"type": "error", "message": "Missing fields: from/to/content."})
                continue

            ok, error_message = send_personal_message_logic(sender, receiver, content, timestamp)
            if not ok:
                logger.warning("WebSocket消息保存失败：%s", error_message)

            send_msg = {
                "type": "message",
                "from": sender,
                "to": receiver,
                "content": content,
                "ts": timestamp,
            }
            await manager.send_to_user(receiver, send_msg)

    except WebSocketDisconnect:
        logger.info("用户 %s 主动断开WebSocket连接", user_id)
        manager.disconnect(user_id)
    except Exception as exc:
        logger.critical("用户 %s 连接发生未预期异常：%s", user_id, exc, exc_info=True)
        manager.disconnect(user_id)
        await websocket.close(code=1011)


if __name__ == "__main__":
    import uvicorn

    uvicorn.run("combined_server:app", host="0.0.0.0", port=8000, reload=True)
