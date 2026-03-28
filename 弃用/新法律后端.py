import logging
import os
from typing import Dict, List

from flask import Flask, jsonify, request, send_from_directory
from openpyxl import load_workbook

BASE_DIR = os.path.abspath(os.path.dirname(__file__))
DATA_PATH = os.path.abspath(os.path.join(BASE_DIR, "..", "数据", "地区.xlsx"))

logging.basicConfig(level=logging.INFO)
logger = logging.getLogger(__name__)

app = Flask(__name__, static_folder=BASE_DIR, static_url_path="/")

_DATA_CACHE: List[Dict[str, str]] = []
_DATA_MTIME: float = -1.0


def _safe_text(value) -> str:
	if value is None:
		return ""
	return str(value).strip()


def _load_excel_rows() -> List[Dict[str, str]]:
	if not os.path.exists(DATA_PATH):
		logger.warning("数据文件不存在: %s", DATA_PATH)
		return []

	workbook = load_workbook(DATA_PATH, read_only=True, data_only=True)
	sheet = workbook.active
	rows: List[Dict[str, str]] = []

	for row in sheet.iter_rows(min_row=2, max_col=3, values_only=True):
		region = _safe_text(row[0])
		url = _safe_text(row[1])
		law_name = _safe_text(row[2])
		if not (region or url or law_name):
			continue
		rows.append({"region": region, "url": url, "name": law_name})

	return rows


def get_data_rows() -> List[Dict[str, str]]:
	global _DATA_CACHE, _DATA_MTIME
	try:
		mtime = os.path.getmtime(DATA_PATH)
	except OSError:
		mtime = -1.0

	if mtime != _DATA_MTIME:
		_DATA_CACHE = _load_excel_rows()
		_DATA_MTIME = mtime
		logger.info("已加载数据: %s 条", len(_DATA_CACHE))

	return _DATA_CACHE


def resolve_location(payload: Dict) -> Dict[str, str]:
	location = payload.get("location") or {}
	province = _safe_text(payload.get("province") or location.get("province"))
	city = _safe_text(payload.get("city") or location.get("city"))
	region = _safe_text(payload.get("region"))

	if region and not province:
		province = region

	if not province:
		province = "全国"

	return {"province": province, "city": city or province}


def build_results(rows: List[Dict[str, str]], keyword: str) -> List[Dict[str, str]]:
	keyword_lower = keyword.lower()
	results: List[Dict[str, str]] = []

	for row in rows:
		name = row.get("name", "")
		region = row.get("region", "")
		if keyword_lower not in name.lower() and keyword_lower not in region.lower():
			continue
		results.append({
			"name": name or "未命名法律",
			"desc": f"地区: {region or '未知'}",
			"url": row.get("url", "")
		})

	return results


def paginate_results(results: List[Dict[str, str]], page: int, page_size: int) -> Dict[str, object]:
	page = max(page, 1)
	page_size = max(min(page_size, 50), 1)
	total = len(results)
	start = (page - 1) * page_size
	end = start + page_size
	return {
		"page": page,
		"page_size": page_size,
		"total": total,
		"results": results[start:end]
	}


def build_recommended(rows: List[Dict[str, str]], location: Dict[str, str]) -> List[Dict[str, str]]:
	province = location.get("province", "")
	city = location.get("city", "")

	if province and province != "全国":
		matched = [
			row for row in rows
			if province in row.get("region", "") or city in row.get("region", "")
		]
	else:
		matched = rows

	recommended: List[Dict[str, str]] = []
	seen = set()
	for row in matched:
		key = (row.get("name", ""), row.get("url", ""))
		if key in seen:
			continue
		seen.add(key)
		recommended.append({
			"title": row.get("name", "未命名法律"),
			"desc": f"来自: {row.get('region', '') or '未知'}"
		})
		if len(recommended) >= 6:
			break

	return recommended


@app.after_request
def add_cors_headers(response):
	response.headers["Access-Control-Allow-Origin"] = "*"
	response.headers["Access-Control-Allow-Headers"] = "Content-Type"
	response.headers["Access-Control-Allow-Methods"] = "POST, OPTIONS, GET"
	return response


@app.route("/search", methods=["POST", "OPTIONS"])
def api_search():
	if request.method == "OPTIONS":
		return ("", 204)

	payload = request.get_json(silent=True) or {}
	keyword = _safe_text(payload.get("keyword"))
	if not keyword:
		return jsonify({"error": "请提供搜索关键字"}), 400

	rows = get_data_rows()
	location = resolve_location(payload)
	all_results = build_results(rows, keyword)
	recommended = build_recommended(rows, location)

	try:
		page = int(payload.get("page", 1))
	except (TypeError, ValueError):
		page = 1
	try:
		page_size = int(payload.get("page_size", 5))
	except (TypeError, ValueError):
		page_size = 5

	page_data = paginate_results(all_results, page, page_size)

	return jsonify({
		"location": location,
		"recommended": recommended,
		"results": page_data["results"],
		"page": page_data["page"],
		"page_size": page_data["page_size"],
		"total": page_data["total"]
	})


@app.route("/location", methods=["POST", "GET", "OPTIONS"])
def api_location():
	if request.method == "OPTIONS":
		return ("", 204)

	if request.method == "GET":
		payload = {}
	else:
		payload = request.get_json(silent=True) or {}

	location = resolve_location(payload)
	return jsonify({"location": location})


@app.route("/")
def index():
	return send_from_directory(app.static_folder, "法律咨询.html")


if __name__ == "__main__":
	app.run(host="0.0.0.0", port=5500, debug=True)
