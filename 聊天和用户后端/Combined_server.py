from contextlib import asynccontextmanager
import threading
import asyncio
import json
import logging
import os
import sys
from logging.handlers import RotatingFileHandler
from typing import Dict, List, Optional, Tuple

import requests
import datetime
from fastapi import FastAPI, Request, WebSocket, WebSocketDisconnect
import uuid
from fastapi.middleware.cors import CORSMiddleware
from fastapi.responses import JSONResponse, PlainTextResponse
from openpyxl import load_workbook
import pandas as pd
from typing import List, Dict, Optional, Any, cast, Union
from openai import AsyncOpenAI
from openai.types.chat import (
    ChatCompletionSystemMessageParam,
    ChatCompletionUserMessageParam,
    ChatCompletionAssistantMessageParam,
)
from dotenv import load_dotenv
load_dotenv()

try:
    from passlib.hash import argon2
except Exception:
    argon2 = None

# On Windows, psycopg async requires a selector-based event loop policy.
# The actual policy will be set after logger is configured to allow logging.


try:
    from postgres_data import adapter as pg_adapter  # type: ignore
    _PG_ADAPTER_AVAILABLE = True
except Exception:
    pg_adapter = None
    _PG_ADAPTER_AVAILABLE = False


# NOTE: `_run_coro_with_selector` 已被移除以避免在线程中运行协程。
# 所有异步 DB 操作应在主事件循环中直接 `await`，或在短期脚本中使用
# postgres_data.adapter 提供的同步 `_sync` 辅助函数。

# ======================================================
# Combined_server.py
# 主服务：处理聊天、用户、群组和帖子管理的 FastAPI 应用。
# - 持久化采用本地 pickle 文件（演示/开发级别）。
# - 将阻塞的模型调用移到线程池以避免阻塞事件循环。
# - 所有磁盘写操作在全局线程锁下进行以避免并发写入冲突。
# 本文件侧重于可读性与可维护性，注释均为中文以方便本地开发阅读。
# ======================================================


# 让BASE_DIR指向上一级目录（主目录）
BASE_DIR = os.path.abspath(os.path.join(os.path.dirname(__file__), '..'))
CHAT_BACKEND_DIR = os.path.dirname(__file__)

# Ensure project root is on sys.path so sibling packages (e.g. postgres_data)
# can be imported when running this file directly (python 聊天和用户后端\Combined_server.py).
if BASE_DIR not in sys.path:
    # insert at front to prefer local packages over site-packages
    sys.path.insert(0, BASE_DIR)

# Keep backend dir on path for local imports within the package
if CHAT_BACKEND_DIR not in sys.path:
    sys.path.append(CHAT_BACKEND_DIR)

from ChatMessage import MessageManage as MsgMgr  # noqa: E402
from ChatMessage import personalChatMessage, groupChatMessage  # noqa: E402
from group import groupManage as GroupMgr  # noqa: E402
from post import PostManage  # noqa: E402
from user import user as UserClass  # noqa: E402
from user import userManage as UserMgr  # noqa: E402
from message_retry import MessageRetryManager  # noqa: E402

# ===================== 日志配置 =====================
LOG_DIR = os.path.join(BASE_DIR, "logs")
os.makedirs(LOG_DIR, exist_ok=True)
LOG_FORMAT = "%(asctime)s - %(levelname)s - %(module)s:%(funcName)s:%(lineno)d - %(message)s"
DATE_FORMAT = "%Y-%m-%d %H:%M:%S"

file_handler = RotatingFileHandler(
    os.path.join(LOG_DIR, "combined_server.log"),
    maxBytes=5 * 1024 * 1024,
    backupCount=5,
    encoding="utf-8",
)
file_handler.setFormatter(logging.Formatter(LOG_FORMAT, DATE_FORMAT))

console_handler = logging.StreamHandler()
console_handler.setFormatter(logging.Formatter(LOG_FORMAT, DATE_FORMAT))

logging.basicConfig(level=logging.INFO, handlers=[file_handler, console_handler])
logger = logging.getLogger(__name__)
# Use default event loop (uvicorn manages the loop). Ensure asyncpg is installed for async DB.
if sys.platform.startswith('win'):
    # 注意：从 Python 3.14 起，asyncio 的 policy 系统（如
    # set_event_loop_policy / WindowsSelectorEventLoopPolicy）已被弃用并将在
    # 未来版本移除。不要在模块导入时设置全局 policy；应在程序入口处
    # 使用 `asyncio.run(..., loop_factory=...)` 或 `asyncio.Runner(loop_factory=...)`。
    logger.info("Windows 平台：不要在模块顶部设置事件循环策略；将在程序入口使用 loop_factory。")
# ===================== 数据路径 =====================
# 数据文件目录（历史上用于本地备份）
BASE_FOLDER = "数据库"

DATA_PATH = os.path.join(BASE_DIR, "数据", "地区.xlsx")
CASES_FILE = os.path.join(BASE_DIR, "数据", "案号.xlsx")

# 配置：是否以数据库为唯一数据源（优先环境变量，默认开启）
DB_ONLY = os.environ.get("DB_ONLY", "1") in ("1", "true", "True")
# 配置：是否启用内存读取缓存（默认关闭）。若关闭，所有读写操作均直接访问 DB。
USE_CACHE = os.environ.get("USE_CACHE", "0") in ("1", "true", "True")


# 日志目录也指向主目录
LOG_DIR = os.path.join(BASE_DIR, "logs")
os.makedirs(LOG_DIR, exist_ok=True)
#==========ai初始化===========  
AIClient = AsyncOpenAI(
    api_key=os.getenv("AI_API_KEY"),  # 从智谱AI开放平台获取
    base_url=os.getenv("AI_API_BASE_URL")
)
MessageParam = Union[
    ChatCompletionSystemMessageParam,
    ChatCompletionUserMessageParam,
    ChatCompletionAssistantMessageParam,
]
# ===================== 业务对象 =====================
@asynccontextmanager
async def lifespan(app: FastAPI):
    """
    使用 FastAPI 的 lifespan 协调应用的启动与关闭。

    - 启动时在后台线程加载所有数据（避免阻塞事件循环）。
    - 关闭时记录 PID 并在后台线程执行一次性保存（调用 save_all_data_on_exit）。
    这样可以更集中地管理进程生命周期与日志，便于在多进程/热重载场景中排查哪个进程执行了保存操作。
    """
    # make clear we will mutate module-level message_save_task
    global message_save_task

    try:
        # 启动时强制使用 Postgres 作为唯一持久化层；若不可用则停止启动。
        if _PG_ADAPTER_AVAILABLE and pg_adapter is not None:
            try:
                counts = await pg_adapter.ensure_seed_data()
                logger.info("Startup: Postgres 种子检查完成 users=%s posts=%s", counts.get('users'), counts.get('posts'))
            except Exception:
                logger.exception("Startup: Postgres 种子检查或种子初始化失败，停止启动")
                # 统一要求 Postgres 可用，停止服务启动以避免使用本地 pkl 回退
                raise RuntimeError("Postgres 不可用或种子检查失败，停止启动")
        else:
            logger.error("Postgres 适配器不可用：服务必须依赖 Postgres 进行持久化，停止启动")
            raise RuntimeError("未检测到 Postgres 适配器，停止启动")

        # 若 pg_adapter 可用，使用它加载必要的运行时缓存（用户、帖子等），不再读取本地 pkl
        try:
            await load_all_data_on_start()
            logger.info("Startup: 从 Postgres 加载运行时缓存完成（PID=%s）", os.getpid())
        except Exception:
            logger.exception("从 Postgres 加载运行时缓存失败，停止启动")
            raise
    except Exception:
        logger.exception("Startup: 加载数据失败")
    # 启动消息重试管理器：在 DB 写失败时持久化并重试发送
    try:
        global message_retry_manager
        retry_file = os.path.join(BASE_DIR, BASE_FOLDER, 'pending_messages.jsonl')
        message_retry_manager = MessageRetryManager(retry_file, retry_interval=5.0)
        await message_retry_manager.start()
    except Exception:
        logger.exception("启动 MessageRetryManager 失败")
    try:
        yield
    finally:
        try:
            pid = os.getpid()
            logger.info("Shutdown: PID=%s 开始保存数据", pid)
            # 停止消息重试管理器
            try:
                if message_retry_manager is not None:
                    await message_retry_manager.stop()
            except Exception:
                logger.exception("停止 MessageRetryManager 失败")

            # 不再进行本地 pkl 的保存；所有持久化以 Postgres 为准
            logger.info("Shutdown: 跳过本地 pkl 保存，Postgres 为唯一持久化层")

            # 尝试优雅关闭 async engine，使用 postgres_data.db_session.dispose_db()
            try:
                from postgres_data import db_session as pg_db_session
                if hasattr(pg_db_session, 'dispose_db'):
                    try:
                        await pg_db_session.dispose_db()
                        logger.info("已优雅关闭 Postgres async engine")
                    except Exception:
                        logger.exception("关闭 async engine 失败")
            except Exception:
                logger.exception("尝试关闭 Postgres 引擎时发生错误")
        except Exception:
            logger.exception("Shutdown: 保存数据失败")


app = FastAPI(lifespan=lifespan)

msg_manager = MsgMgr()
user_manager = UserMgr()
group_manager = GroupMgr()
post_manager = PostManage()
cases_df = pd.DataFrame() # 全局存储案例数据

message_retry_manager: Optional[MessageRetryManager] = None

# 全局写入锁，防止并发写文件
write_lock = threading.Lock()

# 控制本地模型并发调用数：根据机器能力设置为2或3
MODEL_CONCURRENCY = 2
model_semaphore = asyncio.Semaphore(MODEL_CONCURRENCY)


# Helpers: run blocking save operations in threadpool while holding write_lock
def _locked_exec(func, *args, **kwargs):
    """
    在持有模块级 `write_lock`（threading.Lock）的情况下，同步执行
    `func(*args, **kwargs)`。

    该函数设计为在后台线程中运行（例如通过 `asyncio.to_thread`），
    将锁的获取集中在这里可以统一控制对磁盘写入的互斥访问。
    """
    with write_lock:
        return func(*args, **kwargs)

async def run_in_thread_with_lock(func, *args, **kwargs):
    """
    将同步可调用对象下放到后台线程执行，同时保证执行期间持有
    全局的 `write_lock`。

    用法示例：`await run_in_thread_with_lock(some_manager.save, filename)`
    """
    return await asyncio.to_thread(_locked_exec, func, *args, **kwargs)


app.add_middleware(
    CORSMiddleware,
    allow_origins=["*"],
    allow_credentials=True,
    allow_methods=["*"],
    allow_headers=["*"],
)


# 处理 CORS 预检请求（OPTIONS），捕获所有路径的预检请求并返回 200
@app.options("/{path:path}")
async def preflight_handler(path: str, request: Request):
    """
    统一响应浏览器的 CORS 预检请求。CORS 中间件会自动添加必要响应头。
    """
    return JSONResponse(content={}, status_code=200)

# ===================== 工具函数 =====================

_DATA_CACHE: List[Dict[str, str]] = []
_DATA_MTIME: float = -1.0
_save_exit_called = False
# 用于保护数据缓存的异步锁
_data_cache_lock = asyncio.Lock()
def ensure_file_exists(filename: str) -> None:
    """
    确保指定路径的文件存在；如果不存在则创建一个空文件。

    说明：提前创建空文件可以简化后续的加载逻辑，使其无需在每处
    都先判断文件是否存在。
    """
    if not os.path.exists(filename):
        # Create parent dir if necessary
        parent = os.path.dirname(filename)
        if parent and not os.path.exists(parent):
            os.makedirs(parent, exist_ok=True)
        with open(filename, "wb"):
            pass
        logger.info("创建空数据文件：%s", filename)


def return_error(message: str, code: int = 400, request_data: Optional[Dict] = None) -> JSONResponse:
        """
        统一的错误 JSON 响应构造器，供各 API 接口使用。

        - `message`：可读的错误描述。
        - `code`：HTTP 状态码，同时写入返回体的 `code` 字段。
        - `request_data`：可选的请求数据，便于日志和调试。
        """
        error_info = {
            "code": code,
            "error": message,
            "request_data": request_data or {},
        }
        logger.error("接口错误：%s，请求参数：%s", message, error_info["request_data"])
        return JSONResponse(status_code=code, content=error_info)


def return_success(data: Optional[Dict] = None, message: str = "操作成功") -> JSONResponse:
    """
    统一的成功 JSON 响应构造器。
    """
    success_info = {
        "code": 200,
        "message": message,
        "data": data or {},
    }
    logger.info("接口成功：%s，返回数据：%s", message, data)
    return JSONResponse(content=success_info)


async def load_all_data_on_start() -> None:
    """
    在应用启动时以异步方式将必要的运行时缓存从数据库加载到内存。

    - 仅在 Postgres 可用时加载；若不可用，仅记录日志并返回（不抛异常，避免阻塞启动流程）。
    - 该函数改为纯 async，必须在事件循环内直接 await 调用，禁止在后台线程中使用 run_until_complete。
    """
    if not (_PG_ADAPTER_AVAILABLE and pg_adapter is not None):
        logger.error("load_all_data_on_start: Postgres 适配器不可用，无法加载运行时缓存")
        return

    # 若未启用缓存，跳过将 DB 数据加载到内存的过程（所有读取接口请走 DB）。
    if not USE_CACHE:
        logger.info("load_all_data_on_start: 内存缓存加载已禁用（USE_CACHE=false），跳过")
        return

    # 加载用户到 user_manager（同步更新内存缓存以支持管理界面与短期优化）
    try:
        users = await pg_adapter.fetch_users_rows()
    except Exception:
        logger.exception("从 Postgres 加载 users 失败")
        users = []

    user_manager.user_list = []
    for u in users:
        try:
            uid = u.get('id')
            username = u.get('username')
            identity = u.get('identity') or ''
            location = u.get('location') or ''
            role = u.get('role') or identity
            usr = UserClass(uid, username, identity, '', location, role)
            usr.friends = list(u.get('friends') or [])
            usr.state = u.get('state') or 'offline'
            user_manager.add_user(usr)
        except Exception:
            logger.exception("填充 user_manager 时发生异常: %s", u)

    # 加载帖子到 post_manager
    try:
        posts = await pg_adapter.fetch_posts_rows()
    except Exception:
        logger.exception("从 Postgres 加载 posts 失败")
        posts = []

    post_manager.post_list = []
    max_post_id = 0
    max_comment_id = 0
    from post import Post as PostObj, Comment as CommentObj
    for p in posts:
        try:
            pid = p.get('id')
            author = p.get('author')
            title = p.get('title')
            content = p.get('content')
            section = p.get('section')
            time = p.get('time')
            post_obj = PostObj(pid, author, title, content, section, time)
            # 帖子评论
            for c in p.get('comments', []):
                try:
                    cid = c.get('id')
                    comment = CommentObj(cid, pid, c.get('author'), c.get('content'), c.get('time'))
                    post_obj.add_comment(comment)
                    if isinstance(cid, int) and cid > max_comment_id:
                        max_comment_id = cid
                except Exception:
                    logger.exception("构建评论对象失败: %s", c)
            post_manager.post_list.append(post_obj)
            if isinstance(pid, int) and pid > max_post_id:
                max_post_id = pid
        except Exception:
            logger.exception("填充 post_manager 时发生异常: %s", p)

    post_manager.next_post_id = max_post_id + 1 if max_post_id else 1
    post_manager.next_comment_id = max_comment_id + 1 if max_comment_id else 1

    logger.info("从 Postgres 加载运行时缓存完成：users=%d posts=%d", len(user_manager.user_list), len(post_manager.post_list))


def save_all_data_on_exit() -> None:
    """
    将内存中的所有数据保存到磁盘。该函数可被多次调用，但仅第一次
    会执行保存操作（通过 `_save_exit_called` 进行保护）。

    保存操作在 `write_lock` 下执行，以在关闭时提供单一的磁盘写入序列化点。
    注意：DB 写入由各 API 接口在运行时处理，退出时仅保存本地文件以避免连接池问题。
    """
    # 移除本地 pkl 的保存逻辑：Postgres 为唯一持久化层，退出时不再将内存数据写回本地文件。
    logger.info("save_all_data_on_exit: 跳过本地 pkl 保存（Postgres 为唯一持久化层）")


async def _async_update_user_cache(created: dict) -> None:
    """异步（非阻塞）地将新创建的用户写入内存缓存（仅作可选优化）。"""
    try:
        loop = asyncio.get_running_loop()
    except RuntimeError:
        return

    def _sync():
        try:
            from user import user as UserClass  # local import for safety
            usr_obj = UserClass(created.get('id'), created.get('username'), created.get('identity'), '', created.get('location'), created.get('role'))
            usr_obj.friends = list(created.get('friends') or [])
            usr_obj.state = created.get('state') or 'offline'
            try:
                user_manager.add_user(usr_obj)
            except Exception:
                logger.debug('异步更新 user_manager 缓存失败（可忽略）')
        except Exception:
            logger.exception('构建内存 User 对象失败（异步）')

    loop.run_in_executor(None, _sync)


async def _async_update_post_cache(db_post: dict) -> None:
    """异步将新创建的帖子追加到内存缓存（若启用），不影响主流程结果。"""
    try:
        loop = asyncio.get_running_loop()
    except RuntimeError:
        return

    def _sync():
        try:
            from post import Post as PostObj
            p_obj = PostObj(db_post.get('id'), db_post.get('author'), db_post.get('title'), db_post.get('content'), db_post.get('section'), db_post.get('time'))
            try:
                post_manager.post_list.append(p_obj)
                try:
                    _pid_val = db_post.get('id')
                    _pid_int = int(_pid_val) if _pid_val is not None else None
                    if _pid_int is not None:
                        post_manager.next_post_id = max(getattr(post_manager, 'next_post_id', 1), _pid_int + 1)
                except Exception:
                    pass
            except Exception:
                logger.debug('异步更新 post_manager 缓存失败（可忽略）')
        except Exception:
            logger.exception('构建内存 Post 对象失败（异步）')

    loop.run_in_executor(None, _sync)


async def _async_update_post_comment_cache(post_id: int, db_comment: dict) -> None:
    """异步将新评论追加到内存帖子对象（若存在）。"""
    try:
        loop = asyncio.get_running_loop()
    except RuntimeError:
        return

    def _sync():
        try:
            from post import Comment as CommentObj
            post_obj = post_manager.get_post(post_id)
            if post_obj:
                cobj = CommentObj(db_comment.get('id'), post_id, db_comment.get('author'), db_comment.get('content'), db_comment.get('time'))
                try:
                    post_obj.add_comment(cobj)
                    try:
                        _cid_val = db_comment.get('id')
                        _cid_int = int(_cid_val) if _cid_val is not None else None
                        if _cid_int is not None:
                            post_manager.next_comment_id = max(getattr(post_manager, 'next_comment_id', 1), _cid_int + 1)
                    except Exception:
                        pass
                except Exception:
                    logger.debug('异步更新内存帖子评论失败（可忽略）')
        except Exception:
            logger.exception('构建内存 Comment 对象失败（异步）')

    loop.run_in_executor(None, _sync)


async def _async_update_friend_cache(user_id: int, friend_id: int) -> None:
    """异步在内存缓存中为两位用户互相添加好友（仅作优化）。"""
    try:
        loop = asyncio.get_running_loop()
    except RuntimeError:
        return

    def _sync():
        try:
            try:
                mem_a = user_manager.find_user_by_id(user_id)
                mem_b = user_manager.find_user_by_id(friend_id)
            except Exception:
                return
            if mem_a and friend_id not in mem_a.get_friends():
                try:
                    mem_a.add_friend(friend_id)
                except Exception:
                    logger.debug('异步更新内存用户 A 好友列表失败')
            if mem_b and user_id not in mem_b.get_friends():
                try:
                    mem_b.add_friend(user_id)
                except Exception:
                    logger.debug('异步更新内存用户 B 好友列表失败')
        except Exception:
            logger.debug('异步更新好友缓存发生异常（可忽略）')

    loop.run_in_executor(None, _sync)


def _safe_text(value) -> str:
    """将任意值规范为去除首尾空白的字符串；None 返回空字符串。"""
    if value is None:
        return ""
    return str(value).strip()


def _load_excel_rows() -> List[Dict[str, str]]:
    """
    从 `DATA_PATH` 指定的 Excel 文件加载行，并转换为字典列表，
    每项包含 'region'、'url'、'name' 三个字段，空行将被忽略。
    """
    if not os.path.exists(DATA_PATH):
        logger.warning("数据文件不存在: %s", DATA_PATH)
        return []

    workbook = load_workbook(DATA_PATH, read_only=True, data_only=True)
    sheet = workbook.active
    if sheet is None:
        logger.warning("Excel 表格未找到活动工作表: %s", DATA_PATH)
        return []
    rows: List[Dict[str, str]] = []

    for row in sheet.iter_rows(min_row=2, max_col=3, values_only=True):
        region = _safe_text(row[0])
        url = _safe_text(row[1])
        law_name = _safe_text(row[2])
        if not (region or url or law_name):
            continue
        rows.append({"region": region, "url": url, "name": law_name})

    return rows


async def get_data_rows() -> List[Dict[str, str]]:
    global _DATA_CACHE, _DATA_MTIME
    async with _data_cache_lock:
        # 若 Postgres 适配器可用，则直接从数据库读取并返回（不使用文件缓存）
        if _PG_ADAPTER_AVAILABLE and pg_adapter is not None:
            try:
                rows = await pg_adapter.fetch_example_legal_rows()
                logger.info("从 Postgres 加载法规数据，共 %s 条", len(rows))
                return rows
            except Exception:
                logger.exception("调用 pg_adapter.fetch_example_legal_rows 失败，退回到本地 Excel")

        try:
            mtime = os.path.getmtime(DATA_PATH)
        except OSError:
            mtime = -1.0

        if mtime != _DATA_MTIME:
            _DATA_CACHE = _load_excel_rows()
            _DATA_MTIME = mtime
            logger.info("已加载数据: %s 条", len(_DATA_CACHE))
        return _DATA_CACHE


def resolve_location(payload: Dict) -> Dict[str, str]:
    """
    从请求 payload 中解析用户位置信息，支持多种输入形式（顶层字段
    或嵌套的 `location` 字典）。返回包含 `province` 和 `city` 的字典；
    若省份缺失，默认返回 "全国"，城市会回退为省份。
    """
    location = payload.get("location") or {}
    province = _safe_text(payload.get("province") or location.get("province"))
    city = _safe_text(payload.get("city") or location.get("city"))
    region = _safe_text(payload.get("region"))

    if region and not province:
        province = region

    if not province:
        province = "全国"

    return {"province": province, "city": city or province}


def build_results(rows: List[Dict[str, str]], keyword: str) -> List[Dict[str, str]]:
    """
    按关键字过滤 `rows`，关键字可匹配 'name' 或 'region' 字段（不区分大小写），
    返回供 API 使用的标准化结果列表。
    """
    keyword_lower = keyword.lower()
    results: List[Dict[str, str]] = []

    for row in rows:
        name = row.get("name", "")
        region = row.get("region", "")
        if keyword_lower not in name.lower() and keyword_lower not in region.lower():
            continue
        results.append({
            "name": name or "未命名法律",
            "desc": f"地区: {region or '未知'}",
            "url": row.get("url", ""),
        })

    return results


def paginate_results(results: List[Dict[str, str]], page: int, page_size: int) -> Dict[str, object]:
    """简易分页辅助：约束 `page` 和 `page_size` 的合理范围并返回切片结果。"""
    page = max(page, 1)
    page_size = max(min(page_size, 50), 1)
    total = len(results)
    start = (page - 1) * page_size
    end = start + page_size
    return {
        "page": page,
        "page_size": page_size,
        "total": total,
        "results": results[start:end],
    }


def build_recommended(rows: List[Dict[str, str]], location: Dict[str, str]) -> List[Dict[str, str]]: 
    """
    根据 `location` 构建简短的推荐列表。

    通过 (name, url) 去重，最多返回 6 条结果。
    """
    province = location.get("province", "")
    city = location.get("city", "")

    if province and province != "全国":
        matched = [
            row
            for row in rows
            if province in row.get("region", "") or city in row.get("region", "")
        ]
    else:
        matched = rows

    recommended: List[Dict[str, str]] = []
    seen = set()
    for row in matched:
        key = (row.get("name", ""), row.get("url", ""))
        if key in seen:
            continue
        seen.add(key)
        recommended.append({
            "title": row.get("name", "未命名法律"),
            "desc": f"来自: {row.get('region', '') or '未知'}",
        })
        if len(recommended) >= 6:
            break

    return recommended
# 用户标识解析函数：支持 ID 或用户名，优先尝试 ID 查询，失败后回退到用户名查询。    
async def resolve_user_identifier(identifier: str | int) -> tuple[int | None, dict | None]:
    """
    根据用户标识（可以是数字ID或用户名字符串）从数据库查询用户。
    返回 (user_id, user_dict) ，若不存在则返回 (None, None)
    """
    if not identifier:
        return None, None

    # 检查数据库适配器是否可用
    if not _PG_ADAPTER_AVAILABLE or pg_adapter is None:
        logger.error("resolve_user_identifier: 数据库适配器不可用")
        return None, None

    try:
        # 尝试当作用户名查询
        user = await pg_adapter.get_user_by_username(str(identifier))
        if user:
            uid = user.get('id')
            if uid is not None:
                return int(uid), user
    except (ValueError, TypeError):
        pass

    # 当作ID
    uid = int(identifier)  
    user = await pg_adapter.get_user_by_id(uid)
    if user:
            return int(uid), user
    return None, None

async def send_personal_message_logic(sender_name: int, receiver_name: int, content: str, timestamp: str) -> Tuple[bool, str]:
    """
    校验发送者和接收者是否存在，创建 `personalChatMessage` 并添加到
    消息管理器中。消息的持久化由调用方异步触发，以避免在请求或
    WebSocket 主循环中执行阻塞 IO。
    """
    # 统一使用 Postgres 校验用户存在性并将消息加入本地内存缓存（不再回退到本地用户管理器）
    if not (_PG_ADAPTER_AVAILABLE and pg_adapter is not None):
        return False, "数据库不可用"

    try:
        # 支持 id 或 username
        s_id = None
        r_id = None
        try:
            s_id = int(str(sender_name))
        except Exception:
            s_id = None
        try:
            r_id = int(str(receiver_name))
        except Exception:
            r_id = None

        if s_id is not None:
            s_info = await pg_adapter.get_user_by_id(s_id)
        else:
            s_info = await pg_adapter.get_user_by_username(str(sender_name))

        if r_id is not None:
            r_info = await pg_adapter.get_user_by_id(r_id)
        else:
            r_info = await pg_adapter.get_user_by_username(str(receiver_name))

        if not s_info:
            return False, f"发送方「{sender_name}」不存在"
        if not r_info:
            return False, f"接收方「{receiver_name}」不存在"

        # 将消息加入内存缓存用于即时广播（不作本地文件持久化）
        message = personalChatMessage(str(sender_name), str(receiver_name), content, timestamp)
        msg_manager.add_personal_message(message)
        return True, "私信发送成功"
    except Exception:
        logger.exception("send_personal_message_logic: DB 校验路径出错")
        return False, "内部错误"


class ConnectionManager:
    def __init__(self) -> None:
        """管理按 `user_id` 索引的活动 WebSocket 连接。

        说明：此管理器为轻量实现，直接保存 `WebSocket` 对象；当应用
        水平扩展到多进程或多容器时，应替换为共享的 pub/sub 或实时层。
        """
        self.active_connections: Dict[str, WebSocket] = {}

    async def connect(self, user_id: str, websocket: WebSocket) -> None:
        # 接受来自客户端的 WebSocket 连接，并注册以便后续发送消息。
        await websocket.accept()
        self.active_connections[user_id] = websocket
        logger.info("用户 %s 建立WebSocket连接，当前活跃连接数：%s", user_id, len(self.active_connections))

    def disconnect(self, user_id: str) -> None:
        # 从活动连接表中移除用户的 WebSocket。该方法为同步调用，便于在
        # 清理路径中直接调用。
        if user_id in self.active_connections:
            self.active_connections.pop(user_id, None)
            logger.info("用户 %s 断开WebSocket连接，当前活跃连接数：%s", user_id, len(self.active_connections))

    async def send_to_user(self, user_id: str, message: Dict) -> None:
        # 向已连接的用户发送可序列化为 JSON 的消息；若连接不存在则记录日志并返回。
        websocket = self.active_connections.get(user_id)
        if not websocket:
            logger.warning("用户 %s 无活跃WebSocket连接，消息发送失败：%s", user_id, message)
            return
        await websocket.send_text(json.dumps(message, ensure_ascii=False))
        logger.info("向用户 %s 发送消息：%s", user_id, message)


manager = ConnectionManager()


async def _get_payload(request: Request) -> Optional[Dict]:
    try:
        # 先读取原始 body，便于调试和重复读取
        raw = await request.body()
        # 保存原始 body 供上层使用/调试
        try:
            request.state._raw_body = raw
        except Exception:
            pass
        if not raw:
            return None
        try:
            return json.loads(raw.decode('utf-8'))
        except Exception:
            # 无法解析为 JSON
            return None
    except Exception:
        return None
#用户名映射函数





# 使用 FastAPI 生命周期钩子替代模块级的 atexit 注册，以避免在
# 使用自动重载或多进程启动器时出现重复注册/多次保存的问题。


# 全局异常处理器，统一返回友好错误信息
@app.exception_handler(Exception)
async def handle_global_exception(request: Request, exc: Exception):
    # 记录完整异常和请求体（用于诊断）——日志中保留详细信息
    try:
        payload = await _get_payload(request) or dict(request.query_params)
    except Exception:
        payload = {}
    logger.error("Unhandled exception for %s %s; payload=%s", request.method, request.url.path, payload, exc_info=True)

    # 对外返回：通用错误信息且对可能的敏感字段（例如 password）进行脱敏
    def _redact(obj):
        if isinstance(obj, dict):
            out = {}
            for k, v in obj.items():
                if isinstance(k, str) and 'password' in k.lower():
                    out[k] = "<redacted>"
                else:
                    out[k] = _redact(v)
            return out
        if isinstance(obj, list):
            return [_redact(i) for i in obj]
        return obj

    safe_payload = _redact(payload) if payload else {}
    return JSONResponse(
        status_code=500,
        content={"code": 500, "error": "服务器内部错误", "request_data": safe_payload},
    )
# ------------------- API: 获取所有关键词（去重）-------------------
@app.get('/api/keywords')
async def get_keywords():
    # 优先使用 DB 适配器；回退到本地 DataFrame
    try:
        if _PG_ADAPTER_AVAILABLE and pg_adapter is not None:
            rows = await pg_adapter.fetch_relations_rows()
            kws = set()
            for r in rows:
                for k in ('关键词1', '关键词2', '关键词3'):
                    v = (r.get(k) or '').strip()
                    if v:
                        kws.add(v)
            return list(kws)

        if cases_df.empty:
            return []
        keywords = pd.concat([cases_df['关键词1'], cases_df['关键词2'], cases_df['关键词3']])
        unique_keywords = keywords[keywords != ''].unique().tolist()
        return unique_keywords
    except Exception as e:
        logger.exception("Error getting keywords: %s", e)
        return []

# ------------------- API: 获取案例（支持筛选）-------------------
@app.get('/api/cases')
async def get_cases(search: str = "", keyword: str = ""):
    # 优先使用 DB 适配器，返回与原来 DataFrame.to_dict(orient='records') 相同格式的列表
    try:
        rows = []
        if _PG_ADAPTER_AVAILABLE and pg_adapter is not None:
            rows = await pg_adapter.fetch_relations_rows()
        else:
            if cases_df.empty:
                return []
            rows = [r for r in cases_df.to_dict(orient='records')]

        # 1. 按关键词筛选（精确匹配任意一个关键词列）
        if keyword:
            rows = [r for r in rows if (r.get('关键词1') == keyword or r.get('关键词2') == keyword or r.get('关键词3') == keyword)]

        # 2. 按搜索词筛选（在案号或摘要中模糊匹配，大小写不敏感）
        if search:
            s = search.lower()
            def matches(r):
                an = (r.get('案号') or '').lower()
                summary = (r.get('摘要') or '').lower()
                return s in an or s in summary
            rows = [r for r in rows if matches(r)]

        return rows
    except Exception as e:
        logger.exception("Error getting cases: %s", e)
        return []

# 已在文件顶部注册的全局异常处理器将处理所有未捕获异常，
# 避免在文件中重复定义同名处理函数导致覆盖或类型提示冲突。


@app.get("/")
async def root():
    return PlainTextResponse("Combined server is running.")


@app.get("/config")
async def get_config():
    # 返回前端可读取的运行时配置（可根据部署环境修改）
    # 在生产/HTTPS 环境下，前端应使用相对路径以避免混合内容被浏览器阻止
    return JSONResponse(content={"API_BASE": ""})


@app.post("/load_all_data")
async def load_all_data():
    try:
        await load_all_data_on_start()
        return return_success(message="所有数据加载成功")
    except Exception as exc:
        return return_error(f"数据加载失败：{exc}", 500)


@app.post("/save_all_data")
async def save_all_data():
    try:
        save_all_data_on_exit()
        return return_success(message="所有数据保存成功")
    except Exception as exc:
        return return_error(f"数据保存失败：{exc}", 500)


@app.post("/user_register")
async def user_register(request: Request):
    data = await _get_payload(request)
    if not data:
        return return_error("请求参数不能为空，请传入JSON格式数据")
    # 在 DB_ONLY 模式下要求 DB 可用
    if DB_ONLY and not (_PG_ADAPTER_AVAILABLE and pg_adapter is not None):
        return return_error("注册失败：服务被配置为仅使用数据库（DB_ONLY），但数据库不可用", 503)
    userid=data.get("id")
    username = data.get("username")
    identity = data.get("identity")
    password = data.get("password")
    location = data.get("location")
    role = data.get("role") or identity

    missing_fields = []
    if not userid:
        missing_fields.append("用户ID(id)")
    if not username:
        missing_fields.append("用户名(username)")
    if not identity:
        missing_fields.append("身份标识(identity)")
    if not password:
        missing_fields.append("密码(password)")
    if not location:
        missing_fields.append("位置(location)")

    if missing_fields:
        return return_error(f"注册失败：缺少必填参数 - {', '.join(missing_fields)}")

    # 强制使用 Postgres 写入用户，移除本地回退逻辑
    if not (_PG_ADAPTER_AVAILABLE and pg_adapter is not None):
        return return_error("注册失败：数据库不可用，请稍后重试", 503)

    try:
        username_str = str(username)
        userid_int = int(str(userid))
        # 计算 password_hash（若 argon2 可用）以便写入 DB
        password_hash = None
        if argon2 is not None and password:
            try:
                password_hash = argon2.hash(password)
            except Exception:
                logger.exception("生成 password_hash 失败，继续使用明文写入")

        created = await pg_adapter.create_user(userid_int, username_str, password, identity, location, role, friends=[], password_hash=password_hash)
        if created and created.get('id'):
            logger.info("用户已写入 DB：%s", username_str)
            # 异步尝试更新内存缓存（若启用），不依赖其成功
            try:
                asyncio.create_task(_async_update_user_cache(created))
            except Exception:
                logger.debug('注册后调度异步缓存更新失败（可忽略）')
            return return_success(data={"user": created}, message=f"用户「{username_str}」注册成功（存储于DB）")
        else:
            return return_error("注册失败：用户名或用户ID已存在或创建失败", 400)
    except Exception as exc:
        try:
            from postgres_data.adapter import DatabaseError
            if isinstance(exc, DatabaseError):
                logger.exception("数据库错误：%s", exc)
                return return_error("注册失败：数据库内部错误，请稍后重试", 500)
        except Exception:
            pass
        logger.exception("创建用户时发生错误")
        return return_error("注册失败：内部错误，请稍后重试", 500)


@app.post("/user_login")
async def user_login(request: Request):
    data = await _get_payload(request)
    if not data:
        return return_error("请求参数不能为空，请传入JSON格式数据")

    # 强制使用 Postgres 进行登录验证
    if not (_PG_ADAPTER_AVAILABLE and pg_adapter is not None):
        return return_error("登录失败：数据库不可用，请稍后重试", 503)

    userid = data.get("id")
    password = data.get("password")

    if not userid:
        return return_error("登录失败：缺少用户ID(id)参数")
    if not password:
        return return_error("登录失败：缺少密码(password)参数")

    try:
        userid_int = int(str(userid))
    except Exception:
        return return_error("登录失败：用户ID无效", 400)

    try:
        # 先尝试从 DB 获取凭据（password / password_hash）并验证
        creds = await pg_adapter.get_user_credentials(userid_int)
        if not creds:
            return return_error(f"登录失败：用户ID「{userid}」不存在", 401)

        db_pass = creds.get('password')
        db_hash = creds.get('password_hash')
        db_state = creds.get('state')
        # if db_state == 'online':
        #     return return_error(f"登录失败：用户ID「{userid}」已在线", 403)
        if db_hash:
            if argon2 is None or not hasattr(argon2, 'verify'):
                logger.exception("服务器缺少 argon2，无法验证存储的密码哈希")
                return return_error("登录失败：服务器配置错误，无法验证密码", 500)
            try:
                if not argon2.verify(password, db_hash):
                    return return_error(f"登录失败：用户ID「{userid}」的密码错误", 401)
            except Exception:
                logger.exception("argon2 验证异常")
                return return_error(f"登录失败：用户ID「{userid}」的密码错误", 401)
        elif db_pass is not None:
            if str(db_pass) != str(password):
                return return_error(f"登录失败：用户ID「{userid}」的密码错误", 401)
    except Exception:
        logger.exception("登录验证过程中发生错误")
        return return_error("登录失败：验证过程中发生错误，请稍后重试", 500)

    # 如果到达这里，表示凭据验证通过（未抛出异常），返回用户信息并设置在线状态
    try:
        uinfo = await pg_adapter.get_user_by_id(userid_int)
        if not uinfo:
            return return_error(f"登录失败：用户ID「{userid}」不存在", 401)

        # 原子化设置在线状态以避免并发竞态：优先调用 adapter 中的 set_user_state_if_offline
        try:
            if getattr(pg_adapter, 'set_user_state_if_offline', None):
                ok = await pg_adapter.set_user_state_if_offline(user_id=userid_int)
            else:
                # 退回到非原子写法（兼容旧适配器）
                await pg_adapter.set_user_state(user_id=userid_int, state='online')
                ok = True
            if not ok:
                return return_error(f"登录失败：用户ID「{userid}」已在线", 403)
        except Exception:
            logger.exception('登录后同步 DB 状态失败')
            return return_error("登录失败：内部错误，请稍后重试", 500)

        profile = uinfo
        profile.pop('password', None)
        return return_success(data={"user": profile}, message="用户登录成功")
    except Exception as exc:
        try:
            from postgres_data.adapter import DatabaseError
            if isinstance(exc, DatabaseError):
                logger.exception("数据库错误：%s", exc)
                return return_error("登录失败：数据库内部错误，请稍后重试", 500)
        except Exception:
            pass
        logger.exception("获取登录用户信息失败")
        return return_error("登录失败：内部错误，请稍后重试", 500)

@app.post("/user_logout")
async def user_logout(request: Request):
    data = await _get_payload(request) or {}
    # 若请求体为空，尝试从 query params / headers 回退（提高对不同客户端的兼容性）
    if not data:
        # 记录原始 body 与 headers（若 _get_payload 已保存 raw）以便定位客户端来源
        raw_logged = None
        try:
            raw_logged = getattr(request.state, '_raw_body', b'')
        except Exception:
            raw_logged = b''
        try:
            logger.warning('user_logout: 收到空请求体，raw_body=%r, headers=%s', raw_logged, dict(request.headers))
        except Exception:
            logger.warning('user_logout: 收到空请求体，无法记录 raw_body 或 headers')
        try:
            qp = dict(request.query_params)
            hdr_user = request.headers.get('x-username') or request.headers.get('username')
            fallback_username = qp.get('username') or hdr_user
            fallback_id = qp.get('id') or request.headers.get('x-user-id')
            if fallback_username or fallback_id:
                data = {}
                if fallback_username:
                    data['username'] = fallback_username
                if fallback_id:
                    # 尝试转为 int
                    try:
                        data['id'] = int(fallback_id)
                    except Exception:
                        data['id'] = fallback_id
            else:
                logger.warning('user_logout: 请求体为空且无回退参数，将忽略该登出请求')
                return return_success(message='登出请求已忽略（未提供用户信息）')
        except Exception:
            logger.exception('user_logout: 解析回退参数失败，忽略该登出请求')
            return return_success(message='登出请求已忽略（解析失败）')

    username = data.get("username")
    userid = data.get("id")
    if not username and not userid:
        return return_error("登出失败：缺少用户名(username)或用户ID(id)参数")
    # 统一使用 Postgres 更新用户状态（不再使用本地回退逻辑）
    if not (_PG_ADAPTER_AVAILABLE and pg_adapter is not None):
        return return_error("登出失败：数据库不可用，请稍后重试", 503)

    try:
        if userid is not None:
            try:
                uid = int(userid)
            except Exception:
                uid = None
            if uid is not None:
                await pg_adapter.set_user_state(user_id=uid, state='offline')
            else:
                await pg_adapter.set_user_state(username=str(userid), state='offline')
        else:
            await pg_adapter.set_user_state(username=str(username), state='offline')
        return return_success(message=f'{username}登出成功，状态已更新为离线')
    except Exception:
        logger.exception('登出时更新 DB 状态失败')
        return return_error('登出失败：更新数据库状态时出错', 500)


@app.post("/user_friends")
async def user_friends(request: Request):
    data = await _get_payload(request)
    if not data:
        return return_error("请求参数不能为空，请传入JSON格式数据")

    username = data.get("username")
    if not username:
        return return_error("查询失败：缺少用户名(username)参数")

    # 使用 Postgres 查询用户及其好友列表（不再回退到本地实现）
    if not (_PG_ADAPTER_AVAILABLE and pg_adapter is not None):
        return return_error("查询失败：数据库不可用，请稍后重试", 503)

    try:
        db_user = await pg_adapter.get_user_by_username(username)
        if not db_user:
            return return_error(f"查询失败：用户名「{username}」不存在", 404)

        friends_usernames = db_user.get('friends', []) or []
        friends_profiles = []
        for fn in friends_usernames:
            try:
                finfo = await pg_adapter.get_user_by_id(fn)
                if finfo:
                    finfo.pop('password', None)
                    friends_profiles.append(finfo)
                    continue
            except Exception:
                logger.exception("查询好友 %s 时调用 pg_adapter.get_user_by_username 失败", fn)
                continue

        return return_success(data={"friends": friends_profiles}, message=f"查询到用户「{username}」的好友列表")
    except Exception:
        logger.exception("查询好友失败：%s", username)
        return return_error("查询失败：内部错误，请稍后重试", 500)


@app.get("/user_friends")
async def user_friends_get(username: Optional[str] = None):
    """兼容 GET 请求（query 参数 username），始终返回 JSON。"""
    if not username:
        return return_error("查询失败：缺少用户名(username)参数")

    if not (_PG_ADAPTER_AVAILABLE and pg_adapter is not None):
        return return_error("查询失败：数据库不可用，请稍后重试", 503)

    try:
        db_user = await pg_adapter.get_user_by_username(username)
        if not db_user:
            return return_error(f"查询失败：用户名「{username}」不存在", 404)

        friends_usernames = db_user.get('friends', []) or []
        friends_profiles = []
        for fn in friends_usernames:
            try:
                finfo = await pg_adapter.get_user_by_username(fn)
                if finfo:
                    finfo.pop('password', None)
                    friends_profiles.append(finfo)
                    continue
            except Exception:
                logger.exception("查询好友 %s 时调用 pg_adapter.get_user_by_username 失败", fn)
                continue

        return return_success(data={"friends": friends_profiles}, message=f"查询到用户「{username}」的好友列表")
    except Exception:
        logger.exception("查询好友失败（GET）：%s", username)
        return return_error("查询失败：内部错误，请稍后重试", 500)


@app.post("/user_state_search")
async def user_state_search(request: Request):
    data = await _get_payload(request)
    if not data:
        return return_error("请求参数不能为空，请传入JSON格式数据")

    # 在 DB_ONLY 模式下要求 DB 可用
    if DB_ONLY and not (_PG_ADAPTER_AVAILABLE and pg_adapter is not None):
        return return_error("查询失败：服务被配置为仅使用数据库（DB_ONLY），但数据库不可用", 503)

    username = data.get("username")
    userid = data.get("id")
    if not username and not userid:
        return return_error("查询失败：缺少用户名(username)或用户ID(id)参数")

    # 使用 Postgres 查询用户状态（不再回退到本地 user_manager）
    if not (_PG_ADAPTER_AVAILABLE and pg_adapter is not None):
        return return_error("查询失败：数据库不可用，请稍后重试", 503)

    try:
        if username:
            db_user = await pg_adapter.get_user_by_username(username)
            if not db_user:
                return return_error(f"查询失败：用户名「{username}」不存在", 404)
            return return_success(data={"users": [db_user]}, message="查询到用户的状态信息")

        if userid is not None:
            try:
                uid_int = int(userid)
            except Exception:
                uid_int = None
            if uid_int is None:
                return return_error("查询失败：用户ID格式不正确", 400)
            db_user = await pg_adapter.get_user_by_id(uid_int)
            if not db_user:
                return return_error(f"查询失败：用户ID「{userid}」不存在", 404)
            return return_success(data={"users": [db_user]}, message="查询到用户的状态信息")
    except Exception:
        logger.exception("查询用户状态失败")
        return return_error("查询失败：内部错误，请稍后重试", 500)
@app.post("/search_users")
async def search_users(request: Request):
    data = await _get_payload(request)
    if not data:
        return return_error("请求参数不能为空，请传入JSON格式数据")

    username = data.get("username")
    if not username:
        return return_error("查询失败：缺少用户名(username)参数")

    if not (_PG_ADAPTER_AVAILABLE and pg_adapter is not None):
        return return_error("查询失败：数据库不可用，请稍后重试", 503)

    try:
        db_users = await pg_adapter.get_users_by_name(username)
        if not db_users:
            return return_error(f"查询失败：用户名「{username}」不存在", 404)
        for user in db_users:
            user.pop('password', None)
        return return_success(data={"users": db_users}, message="查询到用户的信息")
    except Exception:
        logger.exception("查询用户信息失败")
        return return_error("查询失败：内部错误，请稍后重试", 500)

@app.post("/send_personal_message")
async def send_personal_message(request: Request):
    data = await _get_payload(request)
    if data is None:
        return return_error("请求参数不能为空，请传入JSON格式数据")

    sender = data.get("sender")
    receiver = data.get("receiver")
    content = data.get("content")
    timestamp = data.get("timestamp")

    missing_fields = []
    if sender is None:
        missing_fields.append("发送方(sender)")
    if receiver is None:
        missing_fields.append("接收方(receiver)")
    if content is None:
        missing_fields.append("消息内容(content)")
    if timestamp is None:
        missing_fields.append("时间戳(timestamp)")

    if missing_fields:
        return return_error(f"发送私信失败：缺少必填参数 - {', '.join(missing_fields)}")

    # 类型断言（帮助静态类型检查器并在运行时捕获异常）
    assert isinstance(sender,int) and isinstance(receiver,int) and isinstance(content,str) and isinstance(timestamp,str), "参数类型不正确"
    # 在 DB_ONLY 模式下要求 DB 可用（消息持久化应由 DB 负责）
    if DB_ONLY and not (_PG_ADAPTER_AVAILABLE and pg_adapter is not None):
        return return_error("发送私信失败：服务被配置为仅使用数据库（DB_ONLY），但数据库不可用", 503)

    ok, message = await send_personal_message_logic(sender, receiver, content, timestamp)
    if not ok:
        return return_error(f"发送私信失败：{message}", 404)

    # 若 DB 可用则优先写入 DB；写入失败时尝试入队重试，否则返回 503
    if _PG_ADAPTER_AVAILABLE and pg_adapter is not None:
        try:
            db_row = await pg_adapter.create_personal_message(sender, receiver, content, timestamp)
            return return_success(data={"message": db_row}, message=f"私信发送成功（存储于DB）：从「{sender}」到「{receiver}」")
        except Exception:
            logger.exception("调用 pg_adapter.create_personal_message 失败")
            # 将消息持久化到重试队列，优先使用 message_retry_manager
            try:
                if message_retry_manager is not None:
                    await message_retry_manager.enqueue_personal(sender, receiver, content, timestamp)
                    return return_success(message=f"私信已入队，稍后重试写入DB：从「{sender}」到「{receiver}」")
            except Exception:
                logger.exception("将私信加入重试队列失败")

    # 若执行到此处，表示 DB 不可用或入队失败，返回错误（不再做本地 pkl 保存）
    return return_error("发送私信失败：数据库不可用且无法入队重试，请稍后重试", 503)


@app.post("/send_group_message")
async def send_group_message(request: Request):
    data = await _get_payload(request)
    if not data:
        return return_error("请求参数不能为空，请传入JSON格式数据")

    sender_name = data.get("sender")
    group_name = data.get("group")
    content = data.get("content")
    timestamp = data.get("timestamp")

    missing_fields = []
    if not sender_name:
        missing_fields.append("发送方(sender)")
    if not group_name:
        missing_fields.append("群组名(group)")
    if not content:
        missing_fields.append("消息内容(content)")
    if not timestamp:
        missing_fields.append("时间戳(timestamp)")

    if missing_fields:
        return return_error(f"发送群消息失败：缺少必填参数 - {', '.join(missing_fields)}")

    # 在 DB_ONLY 模式下要求 DB 可用
    if DB_ONLY and not (_PG_ADAPTER_AVAILABLE and pg_adapter is not None):
        return return_error("发送群消息失败：服务被配置为仅使用数据库（DB_ONLY），但数据库不可用", 503)

    # 明确断言为 str，帮助静态类型检查器（并在运行时捕获异常）
    assert isinstance(sender_name, str), "sender_name must be a str"
    assert isinstance(group_name, str), "group_name must be a str"
    assert isinstance(content, str), "content must be a str"
    assert isinstance(timestamp, str), "timestamp must be a str"

    # 确保 Postgres 适配器可用，避免在后续调用时出现 None 成员访问错误
    if not (_PG_ADAPTER_AVAILABLE and pg_adapter is not None):
        return return_error("发送群消息失败：数据库不可用，请稍后重试", 503)

    # 校验发送者仅使用 DB（内存缓存不可作为权威数据源）
    try:
        if str(sender_name).isdigit():
            try:
                ucheck = await pg_adapter.get_user_by_id(int(sender_name))
            except Exception:
                ucheck = None
        else:
            ucheck = await pg_adapter.get_user_by_username(sender_name)
    except Exception:
        logger.exception('发送群消息: 校验发送者时 DB 查询失败')
        return return_error('发送群消息失败：无法校验发送者，请稍后重试', 500)

    if not ucheck:
        return return_error(f"发送群消息失败：发送方「{sender_name}」不存在", 404)

    # NOTE: 暂不校验群组是否存在（若需严格校验，请在 DB 中添加群组表并由 pg_adapter 提供查询）。

    # 若 DB 可用则写入 DB（优先）；写入失败时尝试入队重试，否则返回 503
    if _PG_ADAPTER_AVAILABLE and pg_adapter is not None:
        try:
            db_row = await pg_adapter.create_group_message(group_name, sender_name, content, timestamp)
            # 将消息加入本地内存缓存以便即时广播（不再写本地 pkl）
            try:
                msg = groupChatMessage(sender_name, group_name, content, timestamp)
                msg_manager.add_group_message(msg)
            except Exception:
                logger.exception("将群消息加入本地缓存失败")
            return return_success(data={"message": db_row}, message=f"群消息发送成功（存储于DB）：由「{sender_name}」发送到群「{group_name}」")
        except Exception:
            logger.exception("调用 pg_adapter.create_group_message 失败")
            try:
                if message_retry_manager is not None:
                    await message_retry_manager.enqueue_group(group_name, sender_name, content, timestamp)
                    return return_success(message=f"群消息已入队，稍后重试写入DB：由「{sender_name}」发送到群「{group_name}」")
            except Exception:
                logger.exception("将群消息加入重试队列失败")

    # DB 不可用或入队失败时，返回错误（移除本地 pkl 保存）
    return return_error("发送群消息失败：数据库不可用且无法入队重试，请稍后重试", 503)


@app.post("/get_personal_messages")
async def get_personal_messages(request: Request):
    data = await _get_payload(request)
    if data is None:
        return return_error("请求参数不能为空，请传入JSON格式数据")
    user1 = data.get("user1")
    user2 = data.get("user2")
    if user1 is None or user2 is None:
        return return_error("查询失败：缺少 user1 或 user2 参数")

    # 仅使用 DB 查询私信记录，不再回退到内存历史（内存仅用于实时广播）
    if not (_PG_ADAPTER_AVAILABLE and pg_adapter is not None):
        return return_error("查询失败：数据库不可用，请稍后重试", 503)

    try:
        # 确保传递给 adapter 的为整数 id
        try:
            uid1 = int(user1) if not isinstance(user1, int) else user1
            uid2 = int(user2) if not isinstance(user2, int) else user2
        except Exception:
            return return_error("查询失败：user1 或 user2 格式不正确，应为用户 ID", 400)

        rows = await pg_adapter.fetch_personal_messages(uid1, uid2)
        return return_success(data={"messages": rows}, message=f"查询到{uid1}与{uid2}的历史私聊消息，共{len(rows)}条")
    except Exception:
        logger.exception("使用 pg_adapter.fetch_personal_messages 查询失败")
        return return_error("查询失败：数据库内部错误，请稍后重试", 500)


@app.get("/get_personal_messages")
async def get_personal_messages_get(user1: Optional[int] = None, user2: Optional[int] = None):
    """支持 GET 查询的兼容接口，接受 query 参数 `user1` 和 `user2`。"""
    if user1 is None or user2 is None:
        return return_error("查询失败：缺少 user1 或 user2 参数")

    if not (_PG_ADAPTER_AVAILABLE and pg_adapter is not None):
        return return_error("查询失败：数据库不可用，请稍后重试", 503)

    try:
        rows = await pg_adapter.fetch_personal_messages(int(user1), int(user2))
        return return_success(data={"messages": rows}, message=f"查询到{int(user1)}与{int(user2)}的历史私聊消息，共{len(rows)}条")
    except Exception:
        logger.exception("使用 pg_adapter.fetch_personal_messages 查询失败（GET）")
        return return_error("查询失败：数据库内部错误，请稍后重试", 500)


@app.get("/get_posts")
async def get_posts(request: Request):
    """
    获取帖子列表，支持按板块(section)筛选和按关键词(keyword)搜索。
    :param section: 板块名称（可选，如 'contract', 'owners', 'security', 'public_use'）
    :param keyword: 搜索关键词（可选）
    """
    section = request.query_params.get("section")
    keyword = request.query_params.get("keyword")

    # 强制使用 Postgres 查询帖子列表（不再回退到内存缓存）
    if not (_PG_ADAPTER_AVAILABLE and pg_adapter is not None):
        return return_error("查询帖子失败：数据库不可用，请稍后重试", 503)

    try:
        posts_rows = await pg_adapter.fetch_posts_rows(section)

        # 关键词过滤（基于 title/content）
        if keyword:
            k = str(keyword).lower().strip()
            posts_rows = [p for p in posts_rows if k in p.get('title', '').lower() or k in p.get('content', '').lower()]

        # 时间排序（字符串时间解析）
        def _parse_time_dict(pdict):
            t = pd.to_datetime(pdict.get('time', ''), errors='coerce')
            if pd.isna(t):
                return pd.Timestamp.min
            return t

        posts_rows.sort(key=_parse_time_dict, reverse=True)
        posts_dict = posts_rows

        msg = f"查询成功"
        if section:
            msg += f"，板块「{section}」"
        if keyword:
            msg += f"，关键词「{keyword}」"
        msg += f"，共{len(posts_dict)}条数据"

        return return_success(data={"posts": posts_dict}, message=msg)
    except Exception as exc:
        logger.exception("查询帖子失败：%s", exc)
        return return_error("查询帖子失败：数据库内部错误，请稍后重试", 500)


@app.get("/get_hot_posts")
async def get_hot_posts(limit: int = 8):
    """返回按热度排序的帖子（默认 top N=8）。
    当前热度计算：以评论数为主，时间为次要排序键（评论数降序，时间降序）。
    """
    # 热帖计算仅基于 DB 数据，不回退到内存缓存
    if not (_PG_ADAPTER_AVAILABLE and pg_adapter is not None):
        return return_error("查询热帖失败：数据库不可用，请稍后重试", 503)

    try:
        posts_rows = await pg_adapter.fetch_posts_rows(None)

        def _parse_time(p):
            t = pd.to_datetime(p.get('time', ''), errors='coerce')
            if pd.isna(t):
                return pd.Timestamp.min
            return t

        posts_sorted = sorted(
            posts_rows,
            key=lambda p: (len(p.get('comments', [])), _parse_time(p)),
            reverse=True,
        )
        top = posts_sorted[: max(1, int(limit))]
        posts_dict = top
        return return_success(data={"posts": posts_dict}, message=f"查询到热帖 top {len(posts_dict)}")
    except Exception as exc:
        logger.exception("获取热帖失败：%s", exc)
        return return_error("查询热帖失败：数据库内部错误，请稍后重试", 500)


@app.post("/create_post")
async def create_post(request: Request):
    data = await _get_payload(request)
    if not data:
        return return_error("请求参数不能为空，请传入JSON格式数据")

    author = data.get("author")
    title = data.get("title")
    content = data.get("content")
    section = data.get("section")

    missing_fields = []
    if not author:
        missing_fields.append("作者(author)")
    if not title:
        missing_fields.append("标题(title)")
    if not content:
        missing_fields.append("内容(content)")
    if not section:
        missing_fields.append("板块(section)")

    if missing_fields:
        return return_error(f"创建帖子失败：缺少必填参数 - {', '.join(missing_fields)}")

    # 强制使用 Postgres 进行创建操作
    if not (_PG_ADAPTER_AVAILABLE and pg_adapter is not None):
        return return_error("创建帖子失败：数据库不可用，请稍后重试", 503)

    # 验证作者存在性（仅在 DB 中校验）
    author_name_str = str(author)
    u = None
    try:
        if author_name_str:
            try:
                u = await pg_adapter.get_user_by_username(author_name_str)
            except Exception:
                u = None
    except Exception:
        logger.exception("调用 pg_adapter.get_user_by_username/get_user_by_id 失败")
        return return_error("创建帖子失败：无法验证作者，请稍后重试", 500)

    if not u:
        print(f"作者「{author}」不存在，无法创建帖子")
        return return_error(f"创建帖子失败：作者「{author}」不存在", 404)

    try:
        author_id = u.get('id')
        db_post = await pg_adapter.create_post(author_id, title, content, section)
        if not db_post or not db_post.get('id'):
            logger.error("pg_adapter.create_post 未返回有效结果")
            return return_error("创建帖子失败：数据库未返回有效结果", 500)
        # 异步尝试更新内存缓存（若启用），不阻塞主请求流程
        try:
            asyncio.create_task(_async_update_post_cache(db_post))
        except Exception:
            logger.debug('创建帖子后调度异步缓存更新失败（可忽略）')

        return return_success(
            data={"post": db_post},
            message=f"帖子「{title}」创建成功（ID：{db_post.get('id')}，存储于DB）",
        )
    except Exception as exc:
        try:
            from postgres_data.adapter import DatabaseError
            if isinstance(exc, DatabaseError):
                logger.exception("数据库错误：%s", exc)
                return return_error("创建帖子失败：数据库内部错误，请稍后重试", 500)
        except Exception:
            pass
        logger.exception("调用 pg_adapter.create_post 失败")
        return return_error("创建帖子失败：内部错误，请稍后重试", 500)


@app.get("/get_post_detail")
async def get_post_detail(request: Request):
    post_id_str = request.query_params.get("post_id")
    if not post_id_str:
        return return_error("查询帖子详情失败：缺少帖子ID(post_id)参数", request_data=dict(request.query_params))

    try:
        post_id = int(str(post_id_str))
    except ValueError:
        return return_error(f"查询帖子详情失败：帖子ID「{post_id_str}」不是有效的数字")

    # 仅使用 DB 查询帖子详情（不再回退到内存）
    if not (_PG_ADAPTER_AVAILABLE and pg_adapter is not None):
        return return_error("查询帖子详情失败：数据库不可用，请稍后重试", 503)

    try:
        post_row = await pg_adapter.get_post_by_id(post_id)
        if not post_row:
            return return_error(f"查询帖子详情失败：未找到ID为{post_id}的帖子", 404)
        return return_success(data={"post": post_row}, message=f"查询到ID为{post_id}的帖子详情")
    except Exception:
        logger.exception("从 pg_adapter.get_post_by_id 获取帖子详情失败")
        return return_error("查询帖子详情失败：数据库内部错误，请稍后重试", 500)


@app.post("/add_comment")
async def add_comment(request: Request):
    data = await _get_payload(request)
    if not data:
        return return_error("请求参数不能为空，请传入JSON格式数据")

    post_id_str = data.get("post_id")
    author = data.get("author")
    content = data.get("content")

    missing_fields = []
    if not post_id_str:
        missing_fields.append("帖子ID(post_id)")
    if not author:
        missing_fields.append("评论作者(author)")
    if not content:
        missing_fields.append("评论内容(content)")

    if missing_fields:
        return return_error(f"添加评论失败：缺少必填参数 - {', '.join(missing_fields)}")

    # 在 DB_ONLY 模式下要求 DB 可用
    if DB_ONLY and not (_PG_ADAPTER_AVAILABLE and pg_adapter is not None):
        return return_error("添加评论失败：服务被配置为仅使用数据库（DB_ONLY），但数据库不可用", 503)

    try:
        post_id = int(str(post_id_str))
    except ValueError:
        return return_error(f"添加评论失败：帖子ID「{post_id_str}」不是有效的数字")

    # 强制使用 Postgres 添加评论，移除本地回退逻辑
    if not (_PG_ADAPTER_AVAILABLE and pg_adapter is not None):
        return return_error("添加评论失败：数据库不可用，请稍后重试", 503)

    # 验证作者存在性（仅使用 DB）
    comment_author_name_str = str(author)
    u = None
    try:
        if comment_author_name_str:
            try:
                cid = comment_author_name_str
                u = await pg_adapter.get_user_by_username(cid)
            except Exception:
                u = None
        
    except Exception:
        logger.exception("调用 pg_adapter.get_user_by_username/get_user_by_id 失败")
        return return_error("添加评论失败：无法验证作者，请稍后重试", 500)

    if not u:
        return return_error(f"添加评论失败：作者「{author}」不存在", 404)

    try:
        db_comment = await pg_adapter.add_comment(post_id, author, content)
        if db_comment and db_comment.get('id'):
            # 异步更新内存缓存（若启用），不阻塞请求
            try:
                asyncio.create_task(_async_update_post_comment_cache(post_id, db_comment))
            except Exception:
                logger.debug('添加评论后调度异步缓存更新失败（可忽略）')

            return return_success(
                data={"comment": db_comment},
                message=f"成功为ID{post_id}的帖子添加评论（评论作者：{author}，存储于DB）",
            )
        else:
            logger.error("pg_adapter.add_comment 未返回有效结果")
            return return_error("添加评论失败：数据库未返回有效结果", 500)
    except Exception as exc:
        try:
            from postgres_data.adapter import DatabaseError
            if isinstance(exc, DatabaseError):
                logger.exception("数据库错误：%s", exc)
                return return_error("添加评论失败：数据库内部错误，请稍后重试", 500)
        except Exception:
            pass
        logger.exception("调用 pg_adapter.add_comment 失败")
        return return_error("添加评论失败：内部错误，请稍后重试", 500)

@app.post("/add_friend")
async def add_friend(request: Request):
    data = await _get_payload(request)
    if not data:
        return return_error("请求参数不能为空，请传入JSON格式数据")

    user_id = data.get("user_id")
    friend_id = data.get("friend_id")

    if not user_id:
        return return_error("添加好友失败：缺少用户ID(user_id)参数")
    if not friend_id:
        return return_error("添加好友失败：缺少好友ID(friend_id)参数")

    if user_id == friend_id:
        return return_error(f"添加好友失败：不能添加自己（用户ID：{user_id}）", 400)

    # 使用 DB 校验并添加好友关系（不回退到内存）
    if not (_PG_ADAPTER_AVAILABLE and pg_adapter is not None):
        return return_error("添加好友失败：数据库不可用，请稍后重试", 503)

    # 校验两位用户在 DB 中存在
    try:
        u_a = await pg_adapter.get_user_by_id(user_id)
        u_b = await pg_adapter.get_user_by_id(friend_id)
    except Exception:
        logger.exception('添加好友: 校验用户时 DB 查询失败')
        return return_error('添加好友失败：无法校验用户，请稍后重试', 500)

    if not u_a or not u_b:
        lost_u=user_id if not u_a else friend_id
        logger.warning('添加好友失败：用户「%s」在数据库中不存在', lost_u)
        return return_error('添加好友失败：其中一方用户在数据库中不存在', 404)

    try:
        ua = int(u_a.get('id') or user_id)
        ub = int(u_b.get('id') or friend_id)
        ok = await pg_adapter.add_friend_db(ua, ub)
        if not ok:
            return return_error('添加好友失败：数据库操作未成功', 500)
        # DB 更新成功后，异步尝试更新内存缓存（若启用），不影响主流程
        try:
            asyncio.create_task(_async_update_friend_cache(ua, ub))
        except Exception:
            logger.debug('添加好友后调度异步缓存更新失败（可忽略）')

        return return_success(message=f"添加好友成功（存储于DB）：{u_a.get('username')} ↔ {u_b.get('username')}")
    except Exception:
        logger.exception('调用 pg_adapter.add_friend_db 失败')
        return return_error('添加好友失败：内部错误，请稍后重试', 500)


@app.post("/search")
async def api_search(request: Request):
    payload = await _get_payload(request) or {}
    keyword = _safe_text(payload.get("keyword"))
    if not keyword:
        return return_error("请提供搜索关键字", 400)

    rows = await get_data_rows()
    location = resolve_location(payload)
    all_results = build_results(rows, keyword)
    recommended = build_recommended(rows, location)

    try:
        page = int(payload.get("page", 1))
    except (TypeError, ValueError):
        page = 1
    try:
        page_size = int(payload.get("page_size", 5))
    except (TypeError, ValueError):
        page_size = 5

    page_data = paginate_results(all_results, page, page_size)

    return JSONResponse(
        content={
            "location": location,
            "recommended": recommended,
            "results": page_data["results"],
            "page": page_data["page"],
            "page_size": page_data["page_size"],
            "total": page_data["total"],
        }
    )


@app.post("/search_posts")
async def search_posts(request: Request):
    data = await _get_payload(request) or {}
    keyword = _safe_text(data.get("keyword"))
    if not keyword:
        return return_error("请提供搜索关键字", 400)

    try:
        # 仅使用 Postgres 进行全文（关键词）过滤，不再回退到本地内存列表
        if _PG_ADAPTER_AVAILABLE and pg_adapter is not None:
            posts_rows = await pg_adapter.fetch_posts_rows(None)
            k = str(keyword).lower().strip()
            filtered = [p for p in posts_rows if k in p.get('title', '').lower() or k in p.get('content', '').lower()]
            return return_success(data={"posts": filtered}, message=f"找到{len(filtered)}条相关帖子")
        else:
            return return_error("搜索失败：数据库不可用", 503)
    except Exception as exc:
        logger.exception("search_posts 处理失败：%s", exc)
        return return_error(f"搜索帖子失败：{exc}", 500)


@app.post("/location")
@app.get("/location")
async def api_location(request: Request):
    if request.method == "GET":
        payload = {}
    else:
        payload = await _get_payload(request) or {}

    # 首先尝试使用请求体或 query 中的位置信息
    location = resolve_location(payload)

    # 如果前端未提供任何省市信息，则尝试从请求中提取客户端 IP 并调用第三方 IP 定位服务回退
    def _get_client_ip(req: Request) -> str:
        # 支持常见代理头
        forwarded = req.headers.get("x-forwarded-for") or req.headers.get("X-Forwarded-For")
        if forwarded:
            return forwarded.split(",")[0].strip()
        try:
            client = getattr(req, "client", None)
            host = getattr(client, "host", None)
            return host or ""
        except Exception:
            return ""

    def _resolve_ip_to_location(ip: str) -> dict:
        if not ip:
            return {}
        try:
            # 使用 ip-api 免费接口（字段：status, regionName, city）
            resp = requests.get(f"http://ip-api.com/json/{ip}?fields=status,regionName,city,query", timeout=5)
            data = resp.json()
            if data.get("status") == "success":
                province = data.get("regionName") or ""
                city = data.get("city") or province or ""
                return {"province": province, "city": city}
        except Exception as exc:
            logger.debug("IP 定位请求失败：%s", exc)
        return {}

    # 规范化为字段安全的字典视图，避免类型检查器报错
    payload_dict = payload if isinstance(payload, dict) else {}

    # 判断是否需要 IP 回退：payload 未提供 province/city 且解析后只有默认值
    # 显式取值以避免在静态检查器中对 .get 的不确定性
    province_val = payload_dict["province"] if "province" in payload_dict else None
    city_val = payload_dict["city"] if "city" in payload_dict else None
    location_val = payload_dict["location"] if ("location" in payload_dict and isinstance(payload_dict["location"], dict)) else {}
    location_province = location_val["province"] if (isinstance(location_val, dict) and "province" in location_val) else None
    location_city = location_val["city"] if (isinstance(location_val, dict) and "city" in location_val) else None

    provided_province = bool(province_val or location_province)
    provided_city = bool(city_val or location_city)

    if not provided_province and not provided_city:
        client_ip = _get_client_ip(request)
        logger.info("/location: 尝试使用客户端 IP 回退定位，ip=%s", client_ip)
        ip_location = _resolve_ip_to_location(client_ip)
        if ip_location:
            # 以 IP 定位结果覆盖或补全 location
            location = resolve_location({"province": ip_location.get("province"), "city": ip_location.get("city")})

    return JSONResponse(content={"location": location})
@app.post("/api/newlegal")
async def api_newlegal(request: Request):
    data = await _get_payload(request)
    if not data or "question" not in data:
        return return_error("请输入要咨询的法律问题", 400)
    question = str(data.get("question", "")).strip()
    if not question:
        return return_error("问题不能为空", 400)
    # 系统提示词
    SYSTEM_PROMPT = """
    你是一个AI法律咨询助手，请基于中国法律法规回答用户问题。

    - 引用具体法律条款作为依据
    - 不确定的内容请告知"建议咨询执业律师"
    - 不提供具体的法律行动建议
    """
    messages:List[MessageParam] = [
        {"role": "system", "content": SYSTEM_PROMPT},
    ]
    # 用户提问
    messages.append({"role": "user", "content": question})
    try:
        logger.info("调用法律助手API，问题前30字符: %s", question[:30])
        response = await AIClient.chat.completions.create(
            model="farui-plus",  
            messages=messages,
            temperature=0.3,  # 法律场景温度不宜过高，保持准确性
            max_tokens=1999,
        )
        logger.info("法律助手API调用成功，回答前50字符: %s", str(response.choices[0].message.content)[:50])
        return JSONResponse(content={"answer": response.choices[0].message.content})
    except Exception as e:
        logger.error("调用法律助手API失败: %s", e)
        return JSONResponse(content={"answer": "抱歉，AI法律助手暂时无法回答，请稍后再试。"})

@app.post("/api/legal")
async def legal_chat(request: Request):
    data = await _get_payload(request)
    if not data or "question" not in data:
        return return_error("请输入要咨询的法律问题", 400)

    question = str(data.get("question", "")).strip()
    if not question:
        return return_error("问题不能为空", 400)

    payload = {
        "model": "deepseek-v3.1:671b-cloud",
        "stream": False,
        "messages": [
            {
                "role": "system",
                "content": "你是一名中华人民共和国执业律师，请在回答中**必须**引用具体的法条编号（如《民法典》第23条）。\n如果法律里没有对应条文，请直接回复“暂无相关法条”。\n请使用简洁、正式的法律语言，切勿捏造法条内容。",
            },
            {"role": "user", "content": question},
        ],
    }

    try:
        logger.info("→ 请求 Ollama（受并发限制）: %s...", question[:30])
        # 在异步环境下使用 Semaphore 限制并发，并把阻塞的 requests 调用移到线程池中执行
        async with model_semaphore:
            try:
                resp = await asyncio.wait_for(
                    asyncio.to_thread(
                        lambda: requests.post("http://127.0.0.1:11434/api/chat", json=payload, timeout=30)
                    ),
                    timeout=40,
                )
            except asyncio.TimeoutError:
                logger.exception("❌ 模型调用超时")
                return return_error("模型调用超时，请稍后重试", 504)

        resp.raise_for_status()
        answer = resp.json().get("message", {}).get("content", "暂无相关法条")
        logger.info("✅ 得到答案（前50字符）: %s", answer[:50])
        return JSONResponse(content={"answer": answer})
    except requests.exceptions.ConnectionError:
        logger.exception("❌ 无法连接到 Ollama")
        return return_error("无法连接到 Ollama，请先启动 Ollama", 500)
    except Exception as exc:
        logger.exception("❌ 未知错误")
        return return_error(f"服务器内部错误：{exc}", 500)


@app.get("/health/db")
async def health_db():
    """数据库健康检查：尝试调用 adapter 的 ensure_seed_data（轻量查询）。

    返回 200 表示 DB 可用，返回 503 表示不可用或发生异常。
    """
    if not (_PG_ADAPTER_AVAILABLE and pg_adapter is not None):
        return JSONResponse(status_code=503, content={"code": 503, "db_ok": False, "error": "Postgres adapter unavailable"})
    try:
        counts = await pg_adapter.ensure_seed_data()
        return JSONResponse(status_code=200, content={"code": 200, "db_ok": True, "counts": counts})
    except Exception as exc:
        logger.exception("/health/db 检查失败: %s", exc)
        return JSONResponse(status_code=503, content={"code": 503, "db_ok": False, "error": str(exc)})


@app.websocket("/ws/private")
async def websocket_private_chat(websocket: WebSocket):
    user_id = websocket.query_params.get("user_id")
    if not user_id:
        logger.error("WebSocket连接失败：未提供user_id参数，关闭连接")
        await websocket.close(code=1008)
        return

    await manager.connect(user_id, websocket)
    # WebSocket 建立后将用户状态设为 online（内存与 DB）
    try:
        # 优先使用 DB 查询用户并尽量在内存中找到对应对象以设置状态
        mem_user = None
        if _PG_ADAPTER_AVAILABLE and pg_adapter is not None:
            try:
                try:
                    uid = int(user_id)
                except Exception:
                    uid = None

                if uid is not None:
                    db_user = await pg_adapter.get_user_by_id(uid)
                else:
                    db_user = await pg_adapter.get_user_by_username(user_id)

                if db_user:
                    # 尝试在内存中找到对应用户对象以更新其状态（非必需）
                    try:
                        mem_user = user_manager.find_user(db_user.get('username'))
                    except Exception:
                        mem_user = None
                # 无论是否找到内存对象，都在 DB 中设置 online
                try:
                    if uid is not None:
                        await pg_adapter.set_user_state(user_id=uid, state='online')
                    else:
                        await pg_adapter.set_user_state(username=user_id, state='online')
                except Exception:
                    logger.exception('WebSocket connect: 同步 DB 状态失败')
            except Exception:
                logger.exception('WebSocket connect: DB 查询失败，跳过内存查找')

        # 仅在 DB 可用时设置 online 状态（不回退到内存用户列表）
        if _PG_ADAPTER_AVAILABLE and pg_adapter is not None:
            try:
                try:
                    uid = int(user_id)
                except Exception:
                    uid = None

                if uid is not None:
                    await pg_adapter.set_user_state(user_id=uid, state='online')
                else:
                    await pg_adapter.set_user_state(username=user_id, state='online')
            except Exception:
                logger.exception('WebSocket connect: 同步 DB 状态失败')
        else:
            logger.warning('WebSocket connect: Postgres 不可用，无法设置用户在线状态')
    except Exception:
        logger.exception('WebSocket connect: 更新在线状态失败')

    try:
        while True:
            data = await websocket.receive_text()
            logger.info("收到用户 %s 发送的原始数据：%s", user_id, data)

            try:
                payload = json.loads(data)
            except json.JSONDecodeError as exc:
                logger.error("用户 %s JSON解析失败：%s", user_id, exc)
                await manager.send_to_user(user_id, {"type": "error", "message": "Invalid JSON payload."})
                continue

            sender_raw = payload.get("from")
            receiver_raw = payload.get("to")
            content = payload.get("content")
            timestamp = payload.get("ts")

            # 1. 解析发送者和接收者的真实数据库 ID
            sender, sender_info = await resolve_user_identifier(sender_raw)
            
            receiver, receiver_info = await resolve_user_identifier(receiver_raw)
            
            if not sender or not receiver:
                error_msg = f"用户不存在: sender={sender_raw}, receiver={receiver_raw}"
                logger.error(error_msg)
                await manager.send_to_user(str(sender_raw), {"type": "error", "message": error_msg})
                continue

            missing_fields = []
            if not sender:
                missing_fields.append("from")
            if not receiver:
                missing_fields.append("to")
            if not content:
                missing_fields.append("content")

            if missing_fields:
                error_msg = f"缺失必要字段：{', '.join(missing_fields)}"
                logger.error("用户 %s %s，payload：%s", user_id, error_msg, payload)
                await manager.send_to_user(user_id, {"type": "error", "message": "Missing fields: from/to/content."})
                continue
            # 当存在 Postgres 适配器时，优先使用 DB 完成存在性校验与持久化写入，避免依赖内存用户对象
            if _PG_ADAPTER_AVAILABLE and pg_adapter is not None:
                try:
                    # 校验发送者和接收者在 DB 中是否存在（支持 id 或 username）
                    s_exists = False
                    r_exists = False
                    try:
                        s_id = sender if isinstance(sender, int) else int(str(sender))
                    except Exception:
                        s_id = None
                    try:
                        r_id = receiver if isinstance(receiver, int) else int(str(receiver))
                    except Exception:
                        r_id = None

                    if s_id is not None:
                        s_info = await pg_adapter.get_user_by_id(s_id)
                    else:
                        s_info = await pg_adapter.get_user_by_username(str(sender))
                    if s_info:
                        s_exists = True

                    if r_id is not None:
                        r_info = await pg_adapter.get_user_by_id(r_id)
                    else:
                        r_info = await pg_adapter.get_user_by_username(str(receiver))
                    if r_info:
                        r_exists = True

                    if not (s_exists and r_exists):
                        await manager.send_to_user(str(sender), {"type": "error", "message": "发送失败：发送方或接收方在数据库中不存在"})
                        logger.warning("WebSocket: 发送失败，DB 中缺少用户：%s -> %s", sender, receiver)
                        continue

                    # DB 写入私信记录
                    try:
                        await pg_adapter.create_personal_message(sender, receiver, content, timestamp)
                    except Exception:
                        logger.exception("WebSocket: 调用 pg_adapter.create_personal_message 失败")
                        try:
                            # 优先入队重试管理器
                            if message_retry_manager is not None:
                                await message_retry_manager.enqueue_personal(sender, receiver, content, timestamp)
                                await manager.send_to_user(str(sender), {"type": "info", "message": "消息已入队，稍后重试写入DB"})
                            else:
                                # 无本地 pkl 回退，告知发送者失败
                                await manager.send_to_user(str(sender), {"type": "error", "message": "消息发送失败：数据库不可用且无重试队列"})
                                logger.warning("WebSocket: 数据库写失败且无重试队列，消息丢失")
                                continue
                        except Exception:
                            logger.exception("WebSocket: 将消息加入重试队列失败，无法持久化")
                            await manager.send_to_user(str(sender), {"type": "error", "message": "消息发送失败：无法将消息入队重试"})
                            continue

                except Exception:
                    logger.exception("WebSocket: DB 路径处理消息出错，尝试入队重试")
                    try:
                        if message_retry_manager is not None:
                            await message_retry_manager.enqueue_personal(sender, receiver, content, timestamp)
                            await manager.send_to_user(str(sender), {"type": "info", "message": "消息已入队，稍后重试写入DB"})
                        else:
                            await manager.send_to_user(str(sender), {"type": "error", "message": "消息发送失败：数据库错误且无重试队列"})
                            logger.warning("WebSocket: DB 写失败且无重试队列，消息无法持久化")
                        continue
                    except Exception:
                        logger.exception("WebSocket: 将消息加入重试队列失败")
                        await manager.send_to_user(str(sender), {"type": "error", "message": "消息发送失败：无法将消息入队重试"})
                        continue

            # 若 Postgres 不可用，尝试将消息入队至重试管理器；若也不可用则通知发送者错误
            if not (_PG_ADAPTER_AVAILABLE and pg_adapter is not None):
                try:
                    if message_retry_manager is not None:
                        await message_retry_manager.enqueue_personal(sender, receiver, content, timestamp)
                        await manager.send_to_user(str(sender), {"type": "info", "message": "消息已入队，稍后重试写入DB"})
                    else:
                        await manager.send_to_user(str(sender), {"type": "error", "message": "消息发送失败：数据库不可用且无重试队列"})
                        logger.warning("WebSocket: 数据库不可用且无重试队列，消息无法持久化")
                        continue
                except Exception:
                    logger.exception("WebSocket: 将消息加入重试队列失败，无法持久化")
                    await manager.send_to_user(str(sender), {"type": "error", "message": "消息发送失败：无法将消息入队重试"})
                    continue

            send_msg = {
                "type": "message",
                "from": sender,
                "to": receiver,
                "content": content,
                "ts": timestamp,
            }
            await manager.send_to_user(str(receiver), send_msg)

    except WebSocketDisconnect:
        logger.info("用户 %s 主动断开WebSocket连接", user_id)
        manager.disconnect(user_id)
        # 断开时同步状态为 offline
        try:
            # 仅在 DB 可用时设置 offline（不回退到内存 user_manager）
            if _PG_ADAPTER_AVAILABLE and pg_adapter is not None:
                try:
                    try:
                        uid = int(user_id)
                    except Exception:
                        uid = None
                    if uid is not None:
                        await pg_adapter.set_user_state(user_id=uid, state='offline')
                    else:
                        await pg_adapter.set_user_state(username=user_id, state='offline')
                except Exception:
                    logger.exception('WebSocketDisconnect: 同步 DB 状态失败')
            else:
                logger.warning('WebSocketDisconnect: Postgres 不可用，无法设置用户离线状态')
        except Exception:
            logger.exception('WebSocketDisconnect: 更新离线状态失败')
    except Exception as exc:
        logger.critical("用户 %s 连接发生未预期异常：%s", user_id, exc, exc_info=True)
        manager.disconnect(user_id)
        # 连接异常时同步离线状态
        try:
            if _PG_ADAPTER_AVAILABLE and pg_adapter is not None:
                try:
                    try:
                        uid = int(user_id)
                    except Exception:
                        uid = None
                    if uid is not None:
                        await pg_adapter.set_user_state(user_id=uid, state='offline')
                    else:
                        await pg_adapter.set_user_state(username=user_id, state='offline')
                except Exception:
                    logger.exception('WebSocket exception: 同步 DB 状态失败')
            else:
                logger.warning('WebSocket exception: Postgres 不可用，无法设置用户离线状态')
        except Exception:
            logger.exception('WebSocket exception: 更新离线状态失败')
        await websocket.close(code=1011)


if __name__ == "__main__":
    import uvicorn

    # 以模块方式直接运行时，使用 asyncio.run / asyncio.Runner 启动 uvicorn
    # 并在需要时通过 loop_factory 指定 SelectorEventLoop（避免 Windows 上
    # Proactor 与某些 async DB 驱动的兼容性问题）。保留回退到同步
    # uvicorn.run 以兼容旧环境。
    async def _uvicorn_serve():
        config = uvicorn.Config("Combined_server:app", host="0.0.0.0", port=8000, reload=False)
        server = uvicorn.Server(config)
        await server.serve()

    def _uvicorn_run_sync():
        uvicorn.run("Combined_server:app", host="0.0.0.0", port=8000, reload=False)

    try:
        if sys.platform.startswith("win"):
            # 优先尝试在 asyncio.run 中传入 loop_factory（Python 3.12+）
            try:
                asyncio.run(_uvicorn_serve(), loop_factory=asyncio.SelectorEventLoop)
            except TypeError:
                # loop_factory 参数不可用，尝试使用 asyncio.Runner
                try:
                    with asyncio.Runner(loop_factory=asyncio.SelectorEventLoop) as runner:
                        try:
                            runner.run(_uvicorn_serve())
                        except KeyboardInterrupt:
                            logger.info("入口：收到 KeyboardInterrupt，退出（Runner）")
                        except asyncio.CancelledError:
                            logger.info("入口：Runner 被取消，优雅退出（Runner）")
                except Exception:
                    logger.exception("入口：Runner 启动失败，回退到同步 uvicorn.run")
                    _uvicorn_run_sync()
            except KeyboardInterrupt:
                logger.info("入口：收到 KeyboardInterrupt，退出（asyncio.run loop_factory）")
            except asyncio.CancelledError:
                logger.info("入口：asyncio.run 被取消，优雅退出（Windows loop_factory）")
        else:
            try:
                asyncio.run(_uvicorn_serve())
            except KeyboardInterrupt:
                logger.info("入口：收到 KeyboardInterrupt，退出（asyncio.run）")
            except asyncio.CancelledError:
                logger.info("入口：asyncio.run 被取消，优雅退出")
    except KeyboardInterrupt:
        logger.info("入口：收到 KeyboardInterrupt，退出")
        raise
    except Exception:
        logger.exception("入口：异步启动失败，回退到同步 uvicorn.run")
        _uvicorn_run_sync()
